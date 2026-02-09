import os
import json
import secrets
import traceback
import csv
import io
import base64
import hashlib
import hmac
from functools import wraps
from datetime import datetime

from flask import Flask, render_template, request, jsonify, session, redirect, url_for
from flask import Response
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import text as sql_text
from werkzeug.security import check_password_hash, generate_password_hash

# Optional (recommended) for cross-origin calls from JS
try:
    from flask_cors import CORS
except Exception:
    CORS = None

# Optional Excel reader for bulk uploads
try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

# Optional encryption (required for Aadhaar at rest)
try:
    from cryptography.hazmat.primitives.ciphers.aead import AESGCM
except Exception:
    AESGCM = None

# Optional .env loader (uses current venv if installed)
try:
    from dotenv import load_dotenv
except Exception:
    load_dotenv = None

if load_dotenv:
    load_dotenv()

# ==========================================================
# APP + DATABASE CONFIG
# ==========================================================

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", secrets.token_hex(16))

db = SQLAlchemy()
DB_READY = False  # important guard flag

# ----------------------------------------------------------
# Build DB URL manually from Railway MySQL vars
# ----------------------------------------------------------
mysql_host = os.getenv("MYSQLHOST")
mysql_user = os.getenv("MYSQLUSER")
mysql_password = os.getenv("MYSQLPASSWORD")
mysql_db = os.getenv("MYSQLDATABASE")
mysql_port = os.getenv("MYSQLPORT", "3306")

if mysql_host and mysql_user and mysql_password and mysql_db:
    database_url = (
        f"mysql+pymysql://{mysql_user}:{mysql_password}"
        f"@{mysql_host}:{mysql_port}/{mysql_db}"
    )

    app.config["SQLALCHEMY_DATABASE_URI"] = database_url
    app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
    db.init_app(app)
    DB_READY = True
else:
    print("⚠️ Database environment variables not fully set. DB not initialized.")

# ----------------------------------------------------------
# CORS (unchanged)
# ----------------------------------------------------------
if CORS:
    CORS(app, resources={r"/*": {"origins": "*"}}, supports_credentials=True)


def ensure_questionnaire_tables():
    try:
        def _col_exists(table_name: str, column_name: str) -> bool:
            row = db.session.execute(sql_text("""
                SELECT COUNT(*) AS cnt
                FROM information_schema.columns
                WHERE table_schema = DATABASE()
                  AND table_name = :t
                  AND column_name = :c
            """), {"t": table_name, "c": column_name}).mappings().fetchone()
            return row and int(row["cnt"]) > 0

        ddl = [
            """
            CREATE TABLE IF NOT EXISTS questionnaire_sections (
                section_id INT AUTO_INCREMENT NOT NULL,
                section_order INT NOT NULL,
                section_title VARCHAR(255) NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                PRIMARY KEY (section_id),
                UNIQUE KEY uk_section_order (section_order),
                UNIQUE KEY uk_section_title (section_title),
                INDEX idx_section_order (section_order)
            ) ENGINE=InnoDB;
            """,
            """
            CREATE TABLE IF NOT EXISTS questions (
                question_id INT AUTO_INCREMENT NOT NULL,
                question_order INT NOT NULL,
                question_section_id INT DEFAULT NULL,
                question_text TEXT NOT NULL,
                question_type VARCHAR(50) NOT NULL,
                answer_type VARCHAR(20) NOT NULL,
                options TEXT NULL,
                parent_id INT NULL,
                trigger_value VARCHAR(255) NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                PRIMARY KEY (question_id),
                UNIQUE KEY uk_question_order (question_order),
                INDEX idx_question_order (question_order),
                INDEX idx_question_section_id (question_section_id),
                INDEX idx_questions_parent (parent_id),
                CONSTRAINT fk_question_section
                    FOREIGN KEY (question_section_id)
                    REFERENCES questionnaire_sections(section_id)
                    ON DELETE CASCADE ON UPDATE CASCADE,
                CONSTRAINT fk_parent_question
                    FOREIGN KEY (parent_id)
                    REFERENCES questions(question_id)
                    ON DELETE CASCADE,
                CHECK (question_type IN ('multiple_choice', 'single_choice', 'open_ended')),
                CHECK (answer_type IN ('text', 'numerical'))
            ) ENGINE=InnoDB;
            """,
            """
            CREATE TABLE IF NOT EXISTS individual_questionnaire_sections (
                section_id INT AUTO_INCREMENT NOT NULL,
                section_order INT NOT NULL,
                section_title VARCHAR(255) NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                PRIMARY KEY (section_id),
                UNIQUE KEY uk_individual_section_order (section_order),
                UNIQUE KEY uk_individual_section_title (section_title),
                INDEX idx_individual_section_order (section_order)
            ) ENGINE=InnoDB;
            """,
            """
            CREATE TABLE IF NOT EXISTS individual_questions (
                question_id INT AUTO_INCREMENT NOT NULL,
                question_order INT NOT NULL,
                question_section_id INT DEFAULT NULL,
                question_text TEXT NOT NULL,
                question_type VARCHAR(50) NOT NULL,
                answer_type VARCHAR(20) NOT NULL,
                options TEXT NULL,
                parent_id INT NULL,
                trigger_value VARCHAR(255) NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                PRIMARY KEY (question_id),
                UNIQUE KEY uk_individual_question_order (question_order),
                INDEX idx_individual_question_order (question_order),
                INDEX idx_individual_question_section_id (question_section_id),
                INDEX idx_individual_questions_parent (parent_id),
                CONSTRAINT fk_individual_question_section
                    FOREIGN KEY (question_section_id)
                    REFERENCES individual_questionnaire_sections(section_id)
                    ON DELETE CASCADE ON UPDATE CASCADE,
                CONSTRAINT fk_individual_parent_question
                    FOREIGN KEY (parent_id)
                    REFERENCES individual_questions(question_id)
                    ON DELETE CASCADE,
                CHECK (question_type IN ('multiple_choice', 'single_choice', 'open_ended')),
                CHECK (answer_type IN ('text', 'numerical'))
            ) ENGINE=InnoDB;
            """,
            """
            CREATE TABLE IF NOT EXISTS main_questionnaire_responses (
                main_questionnaire_id BIGINT NOT NULL AUTO_INCREMENT,
                household_id INT NOT NULL,
                user_id INT NOT NULL,
                responses JSON NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (main_questionnaire_id),
                INDEX idx_main_questionnaire_household (household_id),
                INDEX idx_main_questionnaire_user (user_id),
                CONSTRAINT fk_main_questionnaire_household
                    FOREIGN KEY (household_id) REFERENCES households(household_id)
                    ON DELETE CASCADE,
                CONSTRAINT fk_main_questionnaire_user
                    FOREIGN KEY (user_id) REFERENCES users(user_id)
                    ON DELETE CASCADE
            ) ENGINE=InnoDB;
            """,
            """
            CREATE TABLE IF NOT EXISTS survey_contributors (
                contributor_id BIGINT NOT NULL AUTO_INCREMENT,
                main_questionnaire_id BIGINT NOT NULL,
                user_id INT NOT NULL,
                contributed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (contributor_id),
                UNIQUE KEY uk_contributor_main_user (main_questionnaire_id, user_id),
                INDEX idx_contrib_main (main_questionnaire_id),
                INDEX idx_contrib_user (user_id),
                CONSTRAINT fk_contrib_main
                    FOREIGN KEY (main_questionnaire_id)
                    REFERENCES main_questionnaire_responses(main_questionnaire_id)
                    ON DELETE CASCADE,
                CONSTRAINT fk_contrib_user
                    FOREIGN KEY (user_id) REFERENCES users(user_id)
                    ON DELETE CASCADE
            ) ENGINE=InnoDB;
            """,
            """
            CREATE TABLE IF NOT EXISTS individual_questionnaire_responses (
                individual_questionnaire_id BIGINT NOT NULL AUTO_INCREMENT,
                responses JSON NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (individual_questionnaire_id)
            ) ENGINE=InnoDB;
            """,
            """
            CREATE TABLE IF NOT EXISTS main_individual_questionnaire_links (
                link_id BIGINT NOT NULL AUTO_INCREMENT,
                main_questionnaire_id BIGINT NOT NULL,
                individual_questionnaire_id BIGINT NOT NULL,
                household_id INT NOT NULL,
                filled_by_user_id INT NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (link_id),
                INDEX idx_link_main (main_questionnaire_id),
                INDEX idx_link_individual (individual_questionnaire_id),
                INDEX idx_link_household (household_id),
                CONSTRAINT fk_link_main
                    FOREIGN KEY (main_questionnaire_id)
                    REFERENCES main_questionnaire_responses(main_questionnaire_id)
                    ON DELETE CASCADE,
                CONSTRAINT fk_link_individual
                    FOREIGN KEY (individual_questionnaire_id)
                    REFERENCES individual_questionnaire_responses(individual_questionnaire_id)
                    ON DELETE CASCADE,
                CONSTRAINT fk_link_household
                    FOREIGN KEY (household_id) REFERENCES households(household_id)
                    ON DELETE CASCADE,
                CONSTRAINT fk_link_user
                    FOREIGN KEY (filled_by_user_id) REFERENCES users(user_id)
                    ON DELETE CASCADE
            ) ENGINE=InnoDB;
            """
            ,
            """
            CREATE TABLE IF NOT EXISTS household_response_drafts (
                draft_id BIGINT NOT NULL AUTO_INCREMENT,
                household_id INT NOT NULL,
                user_id INT NOT NULL,
                response_data JSON NULL,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                PRIMARY KEY (draft_id),
                UNIQUE KEY uk_draft_household_user (household_id, user_id),
                INDEX idx_draft_household (household_id),
                INDEX idx_draft_user (user_id),
                CONSTRAINT fk_draft_household
                    FOREIGN KEY (household_id) REFERENCES households(household_id)
                    ON DELETE CASCADE,
                CONSTRAINT fk_draft_user
                    FOREIGN KEY (user_id) REFERENCES users(user_id)
                    ON DELETE CASCADE
            ) ENGINE=InnoDB;
            """
        ]

        for stmt in ddl:
            db.session.execute(sql_text(stmt))

        # Backfill missing columns on legacy main questionnaire table
        if not _col_exists("questions", "options"):
            db.session.execute(sql_text("ALTER TABLE questions ADD COLUMN options TEXT NULL"))
        if not _col_exists("questions", "parent_id"):
            db.session.execute(sql_text("ALTER TABLE questions ADD COLUMN parent_id INT NULL"))
        if not _col_exists("questions", "trigger_value"):
            db.session.execute(sql_text("ALTER TABLE questions ADD COLUMN trigger_value VARCHAR(255) NULL"))

        db.session.commit()
    except Exception:
        db.session.rollback()
        print("=== ERROR ENSURING QUESTIONNAIRE TABLES ===")
        print(traceback.format_exc())

def ensure_household_name_unique_index():
    try:
        def _col_exists(table_name: str, column_name: str) -> bool:
            row = db.session.execute(sql_text("""
                SELECT COUNT(*) AS cnt
                FROM information_schema.columns
                WHERE table_schema = DATABASE()
                  AND table_name = :t
                  AND column_name = :c
            """), {"t": table_name, "c": column_name}).mappings().fetchone()
            return row and int(row["cnt"]) > 0

        # Skip if duplicates already exist (would break unique index creation)
        dup = db.session.execute(sql_text("""
            SELECT LOWER(name) AS n, COUNT(*) AS c
            FROM households
            GROUP BY LOWER(name)
            HAVING c > 1
            LIMIT 1
        """)).mappings().fetchone()
        if dup:
            print("=== WARNING: Duplicate household names found; unique index not created ===")
            return

        # Add normalized generated column for case-insensitive uniqueness
        if not _col_exists("households", "name_ci"):
            db.session.execute(sql_text("""
                ALTER TABLE households
                ADD COLUMN name_ci VARCHAR(255)
                GENERATED ALWAYS AS (LOWER(name)) STORED
            """))

        idx = db.session.execute(sql_text("""
            SELECT 1 FROM information_schema.statistics
            WHERE table_schema = DATABASE()
              AND table_name = 'households'
              AND index_name = 'uk_households_name_ci'
            LIMIT 1
        """)).mappings().fetchone()
        if not idx:
            db.session.execute(sql_text("""
                CREATE UNIQUE INDEX uk_households_name_ci ON households (name_ci)
            """))

        db.session.commit()
    except Exception:
        db.session.rollback()
        print("=== ERROR ENSURING HOUSEHOLD UNIQUE INDEX ===")
        print(traceback.format_exc())


def ensure_aadhar_storage_columns():
    try:
        def _col_exists(table_name: str, column_name: str) -> bool:
            row = db.session.execute(sql_text("""
                SELECT COUNT(*) AS cnt
                FROM information_schema.columns
                WHERE table_schema = DATABASE()
                  AND table_name = :t
                  AND column_name = :c
            """), {"t": table_name, "c": column_name}).mappings().fetchone()
            return row and int(row["cnt"]) > 0

        def _idx_exists(table_name: str, index_name: str) -> bool:
            row = db.session.execute(sql_text("""
                SELECT 1
                FROM information_schema.statistics
                WHERE table_schema = DATABASE()
                  AND table_name = :t
                  AND index_name = :i
                LIMIT 1
            """), {"t": table_name, "i": index_name}).mappings().fetchone()
            return bool(row)

        if _col_exists("persons", "aadhar") and not _col_exists("persons", "aadhar_hash"):
            db.session.execute(sql_text("""
                ALTER TABLE persons
                ADD COLUMN aadhar_hash VARCHAR(64) NULL
            """))
        if _col_exists("persons", "aadhar_hash") and not _idx_exists("persons", "idx_persons_aadhar_hash"):
            db.session.execute(sql_text("""
                CREATE INDEX idx_persons_aadhar_hash ON persons (aadhar_hash)
            """))

        if not _col_exists("individual_questionnaire_responses", "aadhar_hash"):
            db.session.execute(sql_text("""
                ALTER TABLE individual_questionnaire_responses
                ADD COLUMN aadhar_hash VARCHAR(64) NULL
            """))
        if _col_exists("individual_questionnaire_responses", "aadhar_hash") and not _idx_exists("individual_questionnaire_responses", "idx_iqr_aadhar_hash"):
            db.session.execute(sql_text("""
                CREATE INDEX idx_iqr_aadhar_hash ON individual_questionnaire_responses (aadhar_hash)
            """))

        db.session.commit()
    except Exception:
        db.session.rollback()
        print("=== ERROR ENSURING AADHAAR STORAGE COLUMNS ===")
        print(traceback.format_exc())


# @app.route("/__init_db__", methods=["POST"])
# def init_db():
#     try:
#         with app.app_context():
#             ensure_questionnaire_tables()
#             ensure_household_name_unique_index()
#             ensure_aadhar_storage_columns()
#         return "✅ Database schema created successfully", 200
#     except Exception:
#         return traceback.format_exc(), 500

# ==========================================================
# DATABASE SCHEMA INITIALIZATION (RAILWAY ONLY)
# ==========================================================
# if DB_READY and os.getenv("RAILWAY_ENVIRONMENT"):
#     # with app.app_context():
#     #     try:
#     #         ensure_questionnaire_tables()
#     #         ensure_household_name_unique_index()
#     #         ensure_aadhar_storage_columns()
#             print("✅ Database schema ensured successfully on Railway")
#         except Exception:
#             print("❌ Database schema initialization failed")
#             print(traceback.format_exc())
# ==========================================================
# DO NOT AUTO-RUN DATABASE MIGRATIONS ON STARTUP
# ==========================================================
print("🚀 App started. Database schema NOT auto-initialized.")

# # ==========================================================
# # APP + DATABASE CONFIG
# # ==========================================================

# app = Flask(__name__)
# app.secret_key = os.environ.get("FLASK_SECRET_KEY", secrets.token_hex(16))

# db = SQLAlchemy()

# # Build DB URL manually from Railway MySQL vars
# mysql_host = os.getenv("MYSQLHOST")
# mysql_user = os.getenv("MYSQLUSER")
# mysql_password = os.getenv("MYSQLPASSWORD")
# mysql_db = os.getenv("MYSQLDATABASE")
# mysql_port = os.getenv("MYSQLPORT", "3306")

# if mysql_host and mysql_user and mysql_password and mysql_db:
#     database_url = (
#         f"mysql+pymysql://{mysql_user}:{mysql_password}"
#         f"@{mysql_host}:{mysql_port}/{mysql_db}"
#     )

#     app.config["SQLALCHEMY_DATABASE_URI"] = database_url
#     app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
#     db.init_app(app)
# if CORS:
#     CORS(app, resources={r"/*": {"origins": "*"}}, supports_credentials=True)

# def ensure_questionnaire_tables():
#     try:
#         def _col_exists(table_name: str, column_name: str) -> bool:
#             row = db.session.execute(sql_text("""
#                 SELECT COUNT(*) AS cnt
#                 FROM information_schema.columns
#                 WHERE table_schema = DATABASE()
#                   AND table_name = :t
#                   AND column_name = :c
#             """), {"t": table_name, "c": column_name}).mappings().fetchone()
#             return row and int(row["cnt"]) > 0

#         ddl = [
#             """
#             CREATE TABLE IF NOT EXISTS questionnaire_sections (
#                 section_id INT AUTO_INCREMENT NOT NULL,
#                 section_order INT NOT NULL,
#                 section_title VARCHAR(255) NOT NULL,
#                 created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
#                 updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
#                 PRIMARY KEY (section_id),
#                 UNIQUE KEY uk_section_order (section_order),
#                 UNIQUE KEY uk_section_title (section_title),
#                 INDEX idx_section_order (section_order)
#             ) ENGINE=InnoDB;
#             """,
#             """
#             CREATE TABLE IF NOT EXISTS questions (
#                 question_id INT AUTO_INCREMENT NOT NULL,
#                 question_order INT NOT NULL,
#                 question_section_id INT DEFAULT NULL,
#                 question_text TEXT NOT NULL,
#                 question_type VARCHAR(50) NOT NULL,
#                 answer_type VARCHAR(20) NOT NULL,
#                 options TEXT NULL,
#                 parent_id INT NULL,
#                 trigger_value VARCHAR(255) NULL,
#                 created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
#                 updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
#                 PRIMARY KEY (question_id),
#                 UNIQUE KEY uk_question_order (question_order),
#                 INDEX idx_question_order (question_order),
#                 INDEX idx_question_section_id (question_section_id),
#                 INDEX idx_questions_parent (parent_id),
#                 CONSTRAINT fk_question_section
#                     FOREIGN KEY (question_section_id)
#                     REFERENCES questionnaire_sections(section_id)
#                     ON DELETE CASCADE ON UPDATE CASCADE,
#                 CONSTRAINT fk_parent_question
#                     FOREIGN KEY (parent_id)
#                     REFERENCES questions(question_id)
#                     ON DELETE CASCADE,
#                 CHECK (question_type IN ('multiple_choice', 'single_choice', 'open_ended')),
#                 CHECK (answer_type IN ('text', 'numerical'))
#             ) ENGINE=InnoDB;
#             """,
#             """
#             CREATE TABLE IF NOT EXISTS individual_questionnaire_sections (
#                 section_id INT AUTO_INCREMENT NOT NULL,
#                 section_order INT NOT NULL,
#                 section_title VARCHAR(255) NOT NULL,
#                 created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
#                 updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
#                 PRIMARY KEY (section_id),
#                 UNIQUE KEY uk_individual_section_order (section_order),
#                 UNIQUE KEY uk_individual_section_title (section_title),
#                 INDEX idx_individual_section_order (section_order)
#             ) ENGINE=InnoDB;
#             """,
#             """
#             CREATE TABLE IF NOT EXISTS individual_questions (
#                 question_id INT AUTO_INCREMENT NOT NULL,
#                 question_order INT NOT NULL,
#                 question_section_id INT DEFAULT NULL,
#                 question_text TEXT NOT NULL,
#                 question_type VARCHAR(50) NOT NULL,
#                 answer_type VARCHAR(20) NOT NULL,
#                 options TEXT NULL,
#                 parent_id INT NULL,
#                 trigger_value VARCHAR(255) NULL,
#                 created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
#                 updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
#                 PRIMARY KEY (question_id),
#                 UNIQUE KEY uk_individual_question_order (question_order),
#                 INDEX idx_individual_question_order (question_order),
#                 INDEX idx_individual_question_section_id (question_section_id),
#                 INDEX idx_individual_questions_parent (parent_id),
#                 CONSTRAINT fk_individual_question_section
#                     FOREIGN KEY (question_section_id)
#                     REFERENCES individual_questionnaire_sections(section_id)
#                     ON DELETE CASCADE ON UPDATE CASCADE,
#                 CONSTRAINT fk_individual_parent_question
#                     FOREIGN KEY (parent_id)
#                     REFERENCES individual_questions(question_id)
#                     ON DELETE CASCADE,
#                 CHECK (question_type IN ('multiple_choice', 'single_choice', 'open_ended')),
#                 CHECK (answer_type IN ('text', 'numerical'))
#             ) ENGINE=InnoDB;
#             """,
#             """
#             CREATE TABLE IF NOT EXISTS main_questionnaire_responses (
#                 main_questionnaire_id BIGINT NOT NULL AUTO_INCREMENT,
#                 household_id INT NOT NULL,
#                 user_id INT NOT NULL,
#                 responses JSON NOT NULL,
#                 created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
#                 PRIMARY KEY (main_questionnaire_id),
#                 INDEX idx_main_questionnaire_household (household_id),
#                 INDEX idx_main_questionnaire_user (user_id),
#                 CONSTRAINT fk_main_questionnaire_household
#                     FOREIGN KEY (household_id) REFERENCES households(household_id)
#                     ON DELETE CASCADE,
#                 CONSTRAINT fk_main_questionnaire_user
#                     FOREIGN KEY (user_id) REFERENCES users(user_id)
#                     ON DELETE CASCADE
#             ) ENGINE=InnoDB;
#             """,
#             """
#             CREATE TABLE IF NOT EXISTS survey_contributors (
#                 contributor_id BIGINT NOT NULL AUTO_INCREMENT,
#                 main_questionnaire_id BIGINT NOT NULL,
#                 user_id INT NOT NULL,
#                 contributed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
#                 PRIMARY KEY (contributor_id),
#                 UNIQUE KEY uk_contributor_main_user (main_questionnaire_id, user_id),
#                 INDEX idx_contrib_main (main_questionnaire_id),
#                 INDEX idx_contrib_user (user_id),
#                 CONSTRAINT fk_contrib_main
#                     FOREIGN KEY (main_questionnaire_id)
#                     REFERENCES main_questionnaire_responses(main_questionnaire_id)
#                     ON DELETE CASCADE,
#                 CONSTRAINT fk_contrib_user
#                     FOREIGN KEY (user_id) REFERENCES users(user_id)
#                     ON DELETE CASCADE
#             ) ENGINE=InnoDB;
#             """,
#             """
#             CREATE TABLE IF NOT EXISTS individual_questionnaire_responses (
#                 individual_questionnaire_id BIGINT NOT NULL AUTO_INCREMENT,
#                 responses JSON NOT NULL,
#                 created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
#                 PRIMARY KEY (individual_questionnaire_id)
#             ) ENGINE=InnoDB;
#             """,
#             """
#             CREATE TABLE IF NOT EXISTS main_individual_questionnaire_links (
#                 link_id BIGINT NOT NULL AUTO_INCREMENT,
#                 main_questionnaire_id BIGINT NOT NULL,
#                 individual_questionnaire_id BIGINT NOT NULL,
#                 household_id INT NOT NULL,
#                 filled_by_user_id INT NOT NULL,
#                 created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
#                 PRIMARY KEY (link_id),
#                 INDEX idx_link_main (main_questionnaire_id),
#                 INDEX idx_link_individual (individual_questionnaire_id),
#                 INDEX idx_link_household (household_id),
#                 CONSTRAINT fk_link_main
#                     FOREIGN KEY (main_questionnaire_id)
#                     REFERENCES main_questionnaire_responses(main_questionnaire_id)
#                     ON DELETE CASCADE,
#                 CONSTRAINT fk_link_individual
#                     FOREIGN KEY (individual_questionnaire_id)
#                     REFERENCES individual_questionnaire_responses(individual_questionnaire_id)
#                     ON DELETE CASCADE,
#                 CONSTRAINT fk_link_household
#                     FOREIGN KEY (household_id) REFERENCES households(household_id)
#                     ON DELETE CASCADE,
#                 CONSTRAINT fk_link_user
#                     FOREIGN KEY (filled_by_user_id) REFERENCES users(user_id)
#                     ON DELETE CASCADE
#             ) ENGINE=InnoDB;
#             """
#             ,
#             """
#             CREATE TABLE IF NOT EXISTS household_response_drafts (
#                 draft_id BIGINT NOT NULL AUTO_INCREMENT,
#                 household_id INT NOT NULL,
#                 user_id INT NOT NULL,
#                 response_data JSON NULL,
#                 updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
#                 PRIMARY KEY (draft_id),
#                 UNIQUE KEY uk_draft_household_user (household_id, user_id),
#                 INDEX idx_draft_household (household_id),
#                 INDEX idx_draft_user (user_id),
#                 CONSTRAINT fk_draft_household
#                     FOREIGN KEY (household_id) REFERENCES households(household_id)
#                     ON DELETE CASCADE,
#                 CONSTRAINT fk_draft_user
#                     FOREIGN KEY (user_id) REFERENCES users(user_id)
#                     ON DELETE CASCADE
#             ) ENGINE=InnoDB;
#             """
#         ]

#         for stmt in ddl:
#             db.session.execute(sql_text(stmt))

#         # Backfill missing columns on legacy main questionnaire table
#         if not _col_exists("questions", "options"):
#             db.session.execute(sql_text("ALTER TABLE questions ADD COLUMN options TEXT NULL"))
#         if not _col_exists("questions", "parent_id"):
#             db.session.execute(sql_text("ALTER TABLE questions ADD COLUMN parent_id INT NULL"))
#         if not _col_exists("questions", "trigger_value"):
#             db.session.execute(sql_text("ALTER TABLE questions ADD COLUMN trigger_value VARCHAR(255) NULL"))

#         db.session.commit()
#     except Exception:
#         db.session.rollback()
#         print("=== ERROR ENSURING QUESTIONNAIRE TABLES ===")
#         print(traceback.format_exc())

# def ensure_household_name_unique_index():
#     try:
#         def _col_exists(table_name: str, column_name: str) -> bool:
#             row = db.session.execute(sql_text("""
#                 SELECT COUNT(*) AS cnt
#                 FROM information_schema.columns
#                 WHERE table_schema = DATABASE()
#                   AND table_name = :t
#                   AND column_name = :c
#             """), {"t": table_name, "c": column_name}).mappings().fetchone()
#             return row and int(row["cnt"]) > 0

#         # Skip if duplicates already exist (would break unique index creation)
#         dup = db.session.execute(sql_text("""
#             SELECT LOWER(name) AS n, COUNT(*) AS c
#             FROM households
#             GROUP BY LOWER(name)
#             HAVING c > 1
#             LIMIT 1
#         """)).mappings().fetchone()
#         if dup:
#             print("=== WARNING: Duplicate household names found; unique index not created ===")
#             return

#         # Add normalized generated column for case-insensitive uniqueness
#         if not _col_exists("households", "name_ci"):
#             db.session.execute(sql_text("""
#                 ALTER TABLE households
#                 ADD COLUMN name_ci VARCHAR(255)
#                 GENERATED ALWAYS AS (LOWER(name)) STORED
#             """))

#         idx = db.session.execute(sql_text("""
#             SELECT 1 FROM information_schema.statistics
#             WHERE table_schema = DATABASE()
#               AND table_name = 'households'
#               AND index_name = 'uk_households_name_ci'
#             LIMIT 1
#         """)).mappings().fetchone()
#         if not idx:
#             db.session.execute(sql_text("""
#                 CREATE UNIQUE INDEX uk_households_name_ci ON households (name_ci)
#             """))

#         db.session.commit()
#     except Exception:
#         db.session.rollback()
#         print("=== ERROR ENSURING HOUSEHOLD UNIQUE INDEX ===")
#         print(traceback.format_exc())


# def ensure_aadhar_storage_columns():
#     try:
#         def _col_exists(table_name: str, column_name: str) -> bool:
#             row = db.session.execute(sql_text("""
#                 SELECT COUNT(*) AS cnt
#                 FROM information_schema.columns
#                 WHERE table_schema = DATABASE()
#                   AND table_name = :t
#                   AND column_name = :c
#             """), {"t": table_name, "c": column_name}).mappings().fetchone()
#             return row and int(row["cnt"]) > 0

#         def _idx_exists(table_name: str, index_name: str) -> bool:
#             row = db.session.execute(sql_text("""
#                 SELECT 1
#                 FROM information_schema.statistics
#                 WHERE table_schema = DATABASE()
#                   AND table_name = :t
#                   AND index_name = :i
#                 LIMIT 1
#             """), {"t": table_name, "i": index_name}).mappings().fetchone()
#             return bool(row)

#         if _col_exists("persons", "aadhar") and not _col_exists("persons", "aadhar_hash"):
#             db.session.execute(sql_text("""
#                 ALTER TABLE persons
#                 ADD COLUMN aadhar_hash VARCHAR(64) NULL
#             """))
#         if _col_exists("persons", "aadhar_hash") and not _idx_exists("persons", "idx_persons_aadhar_hash"):
#             db.session.execute(sql_text("""
#                 CREATE INDEX idx_persons_aadhar_hash ON persons (aadhar_hash)
#             """))

#         if not _col_exists("individual_questionnaire_responses", "aadhar_hash"):
#             db.session.execute(sql_text("""
#                 ALTER TABLE individual_questionnaire_responses
#                 ADD COLUMN aadhar_hash VARCHAR(64) NULL
#             """))
#         if _col_exists("individual_questionnaire_responses", "aadhar_hash") and not _idx_exists("individual_questionnaire_responses", "idx_iqr_aadhar_hash"):
#             db.session.execute(sql_text("""
#                 CREATE INDEX idx_iqr_aadhar_hash ON individual_questionnaire_responses (aadhar_hash)
#             """))

#         db.session.commit()
#     except Exception:
#         db.session.rollback()
#         print("=== ERROR ENSURING AADHAAR STORAGE COLUMNS ===")
#         print(traceback.format_exc())

# with app.app_context():
#     ensure_questionnaire_tables()
#     ensure_household_name_unique_index()
#     ensure_aadhar_storage_columns()


# ==========================================================
# HELPERS / GUARDS
# ==========================================================

def role_required(role: str):
    def deco(f):
        @wraps(f)
        def wrap(*args, **kwargs):
            if session.get("role") != role:
                if role == "user":
                    return redirect(url_for("user_login"))
                if role == "admin":
                    return redirect(url_for("admin_login"))
                return jsonify({"error": "Unauthorized"}), 401
            return f(*args, **kwargs)
        return wrap
    return deco


def json_body():
    return request.get_json(silent=True) or {}


def safe_int(val, default=None):
    try:
        if val is None or val == "":
            return default
        return int(val)
    except Exception:
        return default


def json_error(msg, status=500, details=None):
    payload = {"success": False, "error": msg}
    if details:
        payload["details"] = details
    return jsonify(payload), status


def safe_json_load(val):
    try:
        return json.loads(val) if isinstance(val, str) else val
    except Exception:
        return val


def _extract_status(payload):
    if isinstance(payload, dict):
        return payload.get("status")
    return None


def _combine_draft_sections(draft_payload):
    combined = {}
    for sec in (draft_payload or {}).get("sections", {}).values():
        sec_resp = (sec or {}).get("responses") or {}
        if isinstance(sec_resp, dict):
            combined.update(sec_resp)
    return combined


def _ensure_survey_contributor(main_questionnaire_id, user_id):
    if not main_questionnaire_id or not user_id:
        return
    db.session.execute(sql_text("""
        INSERT INTO survey_contributors (main_questionnaire_id, user_id)
        VALUES (:mid, :uid)
        ON DUPLICATE KEY UPDATE user_id = user_id
    """), {"mid": main_questionnaire_id, "uid": user_id})


def _get_aadhar_enc_key():
    key_raw = (os.environ.get("AADHAR_ENC_KEY") or "").strip()
    if not key_raw:
        raise RuntimeError("AADHAR_ENC_KEY is required for Aadhaar encryption")
    # Accept hex (32 chars) or urlsafe base64
    if len(key_raw) == 32 and all(c in "0123456789abcdefABCDEF" for c in key_raw):
        key = bytes.fromhex(key_raw)
    else:
        key = base64.urlsafe_b64decode(key_raw.encode())
    if len(key) != 16:
        raise RuntimeError("AADHAR_ENC_KEY must decode to 16 bytes (128-bit)")
    return key


def _get_aadhar_hash_key():
    key_raw = (os.environ.get("AADHAR_HASH_KEY") or "").strip()
    if not key_raw:
        # Fall back to encryption key if a dedicated hash key is not provided
        return _get_aadhar_enc_key()
    if len(key_raw) == 64 and all(c in "0123456789abcdefABCDEF" for c in key_raw):
        return bytes.fromhex(key_raw)
    return base64.urlsafe_b64decode(key_raw.encode())


def encrypt_aadhar(plain):
    if plain is None:
        return None
    if AESGCM is None:
        raise RuntimeError("cryptography is required for Aadhaar encryption")
    value = str(plain).strip()
    if not value:
        return ""
    if value.startswith("v1:"):
        return value
    key = _get_aadhar_enc_key()
    aes = AESGCM(key)
    nonce = os.urandom(12)
    ct = aes.encrypt(nonce, value.encode("utf-8"), None)
    token = base64.urlsafe_b64encode(nonce + ct).decode("utf-8")
    return f"v1:{token}"


def decrypt_aadhar(enc):
    if enc is None:
        return None
    value = str(enc)
    if not value:
        return ""
    if not value.startswith("v1:"):
        return value
    if AESGCM is None:
        raise RuntimeError("cryptography is required for Aadhaar decryption")
    key = _get_aadhar_enc_key()
    aes = AESGCM(key)
    token = value[3:]
    raw = base64.urlsafe_b64decode(token.encode("utf-8"))
    nonce, ct = raw[:12], raw[12:]
    plain = aes.decrypt(nonce, ct, None)
    return plain.decode("utf-8")


def hash_aadhar(plain):
    if plain is None:
        return None
    value = "".join(ch for ch in str(plain) if ch.isdigit())
    if not value:
        return ""
    key = _get_aadhar_hash_key()
    return hmac.new(key, value.encode("utf-8"), hashlib.sha256).hexdigest()


def csv_response(filename: str, rows, headers):
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    data = buf.getvalue()
    return Response(
        data,
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def build_question_index(question_table: str, section_table: str):
    rows = db.session.execute(sql_text(f"""
        SELECT q.question_id, q.question_text, q.question_order,
               q.question_section_id, s.section_title, s.section_order
        FROM {question_table} q
        LEFT JOIN {section_table} s ON s.section_id = q.question_section_id
    """)).mappings().all()
    index = {}
    for r in rows:
        index[str(r["question_id"])] = {
            "question_id": r["question_id"],
            "question_text": r["question_text"],
            "question_order": r["question_order"],
            "section_id": r["question_section_id"],
            "section_title": r.get("section_title") or "Unknown Section",
            "section_order": r.get("section_order") or 0,
        }
    return index


def answer_to_str(val):
    if isinstance(val, list):
        return "; ".join(str(v) for v in val)
    if val is None:
        return ""
    return str(val)


def build_export_rows(
    responses_map,
    question_index,
    survey_type,
    household_id=None,
    main_id=None,
    individual_id=None,
    submitted_at=None,
    submitted_by=None,
    member=None,
):
    if not isinstance(responses_map, dict):
        return []
    rows = []
    member = member or {}
    member_name = " ".join([member.get("first_name", ""), member.get("middle_name", ""), member.get("surname", "")]).strip()
    member_aadhar = member.get("aadhar") or ""
    for qid, answer in responses_map.items():
        qmeta = question_index.get(str(qid), {})
        rows.append({
            "survey_type": survey_type,
            "household_id": household_id or "",
            "main_questionnaire_id": main_id or "",
            "individual_questionnaire_id": individual_id or "",
            "member_aadhar": member_aadhar,
            "member_name": member_name,
            "section_order": qmeta.get("section_order", 0),
            "section_title": qmeta.get("section_title", "Unknown Section"),
            "question_order": qmeta.get("question_order", 0),
            "question_id": qmeta.get("question_id", qid),
            "question_text": qmeta.get("question_text", ""),
            "answer": answer_to_str(answer),
            "submitted_at": submitted_at or "",
            "submitted_by": submitted_by or "",
        })
    rows.sort(key=lambda r: (r.get("section_order", 0), r.get("question_order", 0), str(r.get("question_id", ""))))
    return rows


def build_wide_headers(question_index):
    # Use question text with ID to keep headers unique and readable
    ordered = sorted(
        question_index.values(),
        key=lambda r: (r.get("section_order", 0), r.get("question_order", 0), str(r.get("question_id", "")))
    )
    headers = []
    mapping = {}
    used = set()
    for q in ordered:
        qid = q.get("question_id")
        text = (q.get("question_text") or "").strip()
        base = f"{text} [ID:{qid}]" if text else f"Question [ID:{qid}]"
        name = base
        n = 2
        while name in used:
            name = f"{base} ({n})"
            n += 1
        used.add(name)
        headers.append(name)
        mapping[str(qid)] = name
    return headers, mapping


def build_wide_row(responses_map, header_map):
    if not isinstance(responses_map, dict):
        return {}
    row = {}
    for qid, answer in responses_map.items():
        key = header_map.get(str(qid))
        if not key:
            continue
        row[key] = answer_to_str(answer)
    return row


def column_exists(table_name: str, column_name: str) -> bool:
    try:
        row = db.session.execute(sql_text("""
            SELECT COUNT(*) AS cnt
            FROM information_schema.columns
            WHERE table_schema = DATABASE()
              AND table_name = :t
              AND column_name = :c
        """), {"t": table_name, "c": column_name}).mappings().fetchone()
        return row and int(row["cnt"]) > 0
    except Exception:
        return False


def column_is_auto_increment(table_name: str, column_name: str) -> bool:
    try:
        row = db.session.execute(sql_text("""
            SELECT EXTRA AS extra
            FROM information_schema.columns
            WHERE table_schema = DATABASE()
              AND table_name = :t
              AND column_name = :c
        """), {"t": table_name, "c": column_name}).mappings().fetchone()
        extra = (row.get("extra") or "") if row else ""
        return "auto_increment" in extra.lower()
    except Exception:
        return False


def next_id(table_name: str, column_name: str) -> int:
    row = db.session.execute(sql_text(f"""
        SELECT COALESCE(MAX({column_name}), 0) + 1 AS next_id
        FROM {table_name}
    """)).mappings().fetchone()
    return int(row["next_id"]) if row and "next_id" in row else 1


def normalize_header_name(val: str) -> str:
    raw = str(val or "").strip().lower()
    return "".join(ch for ch in raw if ch.isalnum())


def read_excel_rows(file_storage):
    if not load_workbook:
        raise RuntimeError("openpyxl not installed")
    wb = load_workbook(file_storage, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        rows.append([cell if cell is not None else "" for cell in row])
    return rows


def parse_bulk_rows(rows, level: str):
    header_map = {
        "state": "state",
        "states": "state",
        "statename": "state",
        "district": "district",
        "districts": "district",
        "districtname": "district",
        "block": "block",
        "blocks": "block",
        "blockname": "block",
        "subcenter": "subcenter",
        "subcentre": "subcenter",
        "subcenters": "subcenter",
        "subcentres": "subcenter",
        "subcentername": "subcenter",
        "subcentrename": "subcenter",
        "village": "village",
        "villages": "village",
        "villagename": "village",
        "lgd": "lgd",
        "lgdcode": "lgd",
        "villagelgd": "lgd",
        "villagelgdcode": "lgd",
        "villagelgdcodeid": "lgd",
        "villagelgd": "lgd",
        "village_lgd_code": "lgd",
    }

    def has_header(first_row):
        for cell in first_row:
            if header_map.get(normalize_header_name(cell)):
                return True
        return False

    def cell_to_str(v):
        if v is None:
            return ""
        if isinstance(v, bool):
            return "1" if v else "0"
        if isinstance(v, float) and v.is_integer():
            v = int(v)
        return str(v).strip()

    header = has_header(rows[0])
    col_idx = {}
    start_idx = 0
    if header:
        start_idx = 1
        for i, cell in enumerate(rows[0]):
            key = header_map.get(normalize_header_name(cell))
            if key and key not in col_idx:
                col_idx[key] = i

    def get_val(row, key, pos):
        idx = col_idx.get(key, pos)
        if idx is None or idx >= len(row):
            return ""
        return cell_to_str(row[idx])

    required_by_level = {
        "states": ["state"],
        "districts": ["state", "district"],
        "blocks": ["state", "district", "block"],
        "subcenters": ["state", "district", "block", "subcenter"],
        "villages": ["state", "district", "block", "subcenter", "village"],
    }
    positions = {
        "state": 0,
        "district": 1,
        "block": 2,
        "subcenter": 3,
        "village": 4,
        "lgd": 5,
    }

    items = []
    errors = []
    for i in range(start_idx, len(rows)):
        row = rows[i]
        if not any(cell_to_str(c) for c in row):
            continue
        item = {"_row": i + 1}
        for key in ["state", "district", "block", "subcenter", "village", "lgd"]:
            item[key] = get_val(row, key, positions[key])
        missing = [k for k in required_by_level[level] if not item.get(k)]
        if missing:
            errors.append(f"Row {i + 1}: missing {', '.join(missing)}")
            continue
        items.append(item)

    return items, errors


def normalize_question_payload(data: dict):
    """
    Accept multiple payload styles from different UIs:
    - admin UI might send: {section_id, question_text, answer_type, options, parent_id, trigger_value, question_type}
    - old UI might send: {section_id, text, type}
    """

    section_id = safe_int(data.get("section_id"))
    question_text_value = (data.get("question_text") or data.get("text") or "").strip()

    # "answer_type" from frontend might be: number/text/numerical/open_ended/mcq/checkbox
    incoming_type = (data.get("answer_type") or data.get("type") or data.get("question_type") or "").strip()

    # question_type is REQUIRED by DB (example: single_choice, multiple_choice, open_ended)
    # If frontend doesn't send question_type explicitly, we infer.
    # You can adjust mapping according to your UI.
    mapping = {
        "text":            {"question_type": "open_ended",     "answer_type": "text"},
        "open_ended":      {"question_type": "open_ended",     "answer_type": "text"},
        "number":          {"question_type": "open_ended",     "answer_type": "numerical"},
        "numerical":       {"question_type": "open_ended",     "answer_type": "numerical"},
        "mcq":             {"question_type": "single_choice",  "answer_type": "numerical"},
        "single_choice":   {"question_type": "single_choice",  "answer_type": "numerical"},
        "checkbox":        {"question_type": "multiple_choice","answer_type": "numerical"},
        "multiple_choice": {"question_type": "multiple_choice","answer_type": "numerical"},
    }

    inferred = mapping.get(incoming_type, {"question_type": "open_ended", "answer_type": "text"})

    question_type = (data.get("question_type") or "").strip() or inferred["question_type"]
    answer_type = inferred["answer_type"]

    # Options can be list, JSON string, CSV string, or None
    options = data.get("options", None)
    if isinstance(options, list):
        options = json.dumps(options)
    elif isinstance(options, dict):
        options = json.dumps(options)
    elif isinstance(options, str):
        options = options.strip()
        if options == "":
            options = None
        else:
            try:
                parsed = json.loads(options)
                if isinstance(parsed, (list, dict)):
                    options = json.dumps(parsed)
                else:
                    options = json.dumps([str(parsed)])
            except Exception:
                options = json.dumps([o.strip() for o in options.split(",") if o.strip()])

    parent_id = data.get("parent_id", None)
    parent_id = safe_int(parent_id, None)
    if parent_id == 0:
        parent_id = None

    trigger_value = data.get("trigger_value", None)
    if isinstance(trigger_value, str):
        trigger_value = trigger_value.strip()
        if trigger_value == "":
            trigger_value = None

    # Only choice questions should keep options
    if question_type == "open_ended":
        options = None

    return {
        "section_id": section_id,
        "question_text": question_text_value,
        "question_type": question_type,
        "answer_type": answer_type,
        "options": options,
        "parent_id": parent_id,
        "trigger_value": trigger_value
    }

def build_question_tree(questions):
    """
    Build a nested question tree from a flat list of question dicts.
    Expects keys: question_id, parent_id, question_order.
    """
    by_parent = {}
    for q in questions:
        parent_id = q.get("parent_id")
        if parent_id in (0, "0"):
            parent_id = None
            q["parent_id"] = None
        by_parent.setdefault(parent_id, []).append(q)

    for parent_id, items in by_parent.items():
        items.sort(key=lambda r: r.get("question_order") or 0)

    def build(parent_id=None):
        nodes = []
        for q in by_parent.get(parent_id, []):
            node = dict(q)
            node["children"] = build(q["question_id"])
            nodes.append(node)
        return nodes

    return build(None)


# ==========================================================
# PAGES
# ==========================================================

@app.route("/")
def home():
    return render_template("home.html")


@app.route("/user-login", methods=["GET", "POST"])
def user_login():
    if request.method == "POST":
        try:
            result = db.session.execute(
                sql_text("SELECT user_id, password_hash FROM users WHERE username=:u"),
                {"u": request.form["username"]},
            ).mappings().fetchone()

            if result and check_password_hash(result["password_hash"], request.form["password"]):
                session.clear()
                session["user_id"] = result["user_id"]
                session["role"] = "user"
                session["username"] = request.form["username"]
                return redirect(url_for("user_dashboard"))

            return render_template("user_login.html", error="Invalid username or password")
        except Exception as e:
            return render_template("user_login.html", error=f"Login error: {str(e)}")

    return render_template("user_login.html")


@app.route("/admin-login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        try:
            result = db.session.execute(
                sql_text("SELECT admin_id, password_hash FROM admins WHERE username=:u"),
                {"u": request.form["username"]},
            ).mappings().fetchone()

            if result and check_password_hash(result["password_hash"], request.form["password"]):
                session.clear()
                session["admin_id"] = result["admin_id"]
                session["role"] = "admin"
                session["username"] = request.form["username"]
                return redirect(url_for("admin_dashboard"))

            return render_template("admin_login.html", error="Invalid username or password")
        except Exception as e:
            return render_template("admin_login.html", error=f"Login error: {str(e)}")

    return render_template("admin_login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("home"))


@app.route("/user-dashboard")
@role_required("user")
def user_dashboard():
    return render_template("user_dashboard.html")


@app.route("/admin-dashboard")
@role_required("admin")
def admin_dashboard():
    return render_template("admin_dashboard.html")


@app.route("/questionnaire")
@role_required("user")
def questionnaire():
    if "household_id" not in session:
        return redirect(url_for("user_dashboard"))
    return render_template("questionnaire.html")


@app.route("/survey")
@role_required("user")
def survey():
    if "household_id" not in session:
        return redirect(url_for("user_dashboard"))
    return render_template("survey.html")


# ==========================================================
# USER APIs
# ==========================================================

@app.route("/api/current-user")
def get_current_user():
    return jsonify({"username": session.get("username", "User")})

@app.route("/api/current-household")
def get_current_household():
    return jsonify({
        "household_id": session.get("household_id"),
        "main_questionnaire_id": session.get("main_questionnaire_id")
    })


@app.route("/api/household-draft")
def get_household_draft():
    if session.get("role") != "user" or "household_id" not in session:
        return json_error("Unauthorized", 401)

    try:
        household_id = session.get("household_id")
        main_row = db.session.execute(sql_text("""
            SELECT main_questionnaire_id, responses
            FROM main_questionnaire_responses
            WHERE household_id = :hid
            ORDER BY main_questionnaire_id DESC
            LIMIT 1
        """), {"hid": household_id}).mappings().fetchone()

        main_id = main_row["main_questionnaire_id"] if main_row else None
        status = _extract_status(safe_json_load(main_row["responses"])) if main_row else None

        draft_row = db.session.execute(sql_text("""
            SELECT response_data
            FROM household_response_drafts
            WHERE household_id = :hid
            ORDER BY updated_at DESC
            LIMIT 1
        """), {"hid": household_id}).mappings().fetchone()
        draft_payload = safe_json_load(draft_row["response_data"]) if draft_row else {}
        combined = _combine_draft_sections(draft_payload)

        return jsonify({
            "success": True,
            "household_id": household_id,
            "main_questionnaire_id": main_id,
            "status": status,
            "draft": draft_payload,
            "responses": combined
        })
    except Exception as e:
        return json_error(str(e), 500)


@app.route("/api/survey-draft", methods=["POST"])
def save_survey_draft():
    if session.get("role") != "user" or "household_id" not in session:
        return json_error("Unauthorized", 401)

    try:
        data = json_body()
        survey_state = data.get("survey_state")
        if not isinstance(survey_state, dict):
            return jsonify({"success": True})

        household_id = session.get("household_id")
        user_id = session.get("user_id")
        row = db.session.execute(sql_text("""
            SELECT draft_id, response_data
            FROM household_response_drafts
            WHERE household_id = :hid
            ORDER BY updated_at DESC
            LIMIT 1
        """), {"hid": household_id}).mappings().fetchone()

        existing = safe_json_load(row["response_data"]) if row else {}
        if not isinstance(existing, dict):
            existing = {}
        existing["survey_state"] = survey_state
        existing["updated_at"] = datetime.now().isoformat()

        if row and row.get("draft_id"):
            db.session.execute(sql_text("""
                UPDATE household_response_drafts
                SET response_data = :data
                WHERE draft_id = :did
            """), {"did": row["draft_id"], "data": json.dumps(existing)})
        else:
            db.session.execute(sql_text("""
                INSERT INTO household_response_drafts (household_id, user_id, response_data)
                VALUES (:hid, :uid, :data)
            """), {"hid": household_id, "uid": user_id, "data": json.dumps(existing)})

        main_id = safe_int(session.get("main_questionnaire_id"))
        if main_id:
            _ensure_survey_contributor(main_id, user_id)

        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return json_error(str(e), 500)


@app.route("/api/states")
def get_states():
    try:
        rows = db.session.execute(sql_text("SELECT * FROM states ORDER BY name")).mappings().all()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        return json_error(str(e), 500)


@app.route("/api/districts/<int:state_id>")
def get_districts(state_id):
    try:
        rows = db.session.execute(
            sql_text("SELECT * FROM districts WHERE state_id=:sid ORDER BY name"),
            {"sid": state_id},
        ).mappings().all()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        return json_error(str(e), 500)


@app.route("/api/blocks/<int:district_id>")
def get_blocks(district_id):
    try:
        rows = db.session.execute(
            sql_text("SELECT * FROM blocks WHERE district_id=:did ORDER BY name"),
            {"did": district_id},
        ).mappings().all()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        return json_error(str(e), 500)


@app.route("/api/sub-centers/<int:block_id>")
def get_sub_centers(block_id):
    try:
        rows = db.session.execute(
            sql_text("SELECT * FROM sub_centers WHERE block_id=:bid ORDER BY name"),
            {"bid": block_id},
        ).mappings().all()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        return json_error(str(e), 500)


@app.route("/api/villages/by-subcenter/<int:sub_center_id>")
def get_villages_by_subcenter(sub_center_id):
    try:
        rows = db.session.execute(
            sql_text("SELECT village_id, name FROM villages WHERE sub_center_id=:sid ORDER BY name"),
            {"sid": sub_center_id},
        ).mappings().all()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        return json_error(str(e), 500)


@app.route("/api/household", methods=["POST"])
def household():
    if session.get("role") != "user":
        return json_error("Unauthorized", 401)

    try:
        data = json_body()

        required = ["household_name", "state_id", "district_id", "block_id", "sub_center_id", "village_id"]
        for k in required:
            if data.get(k) in (None, ""):
                return json_error(f"{k.replace('_', ' ').title()} is required", 400)

        existing = db.session.execute(
            sql_text("SELECT household_id FROM households WHERE LOWER(name) = LOWER(:name) LIMIT 1"),
            {"name": data["household_name"]}
        ).mappings().fetchone()
        if existing:
            household_id = existing["household_id"]
            main_row = db.session.execute(sql_text("""
                SELECT main_questionnaire_id, user_id, responses
                FROM main_questionnaire_responses
                WHERE household_id = :hid
                ORDER BY main_questionnaire_id DESC
                LIMIT 1
            """), {"hid": household_id}).mappings().fetchone()

            if main_row:
                resp_payload = safe_json_load(main_row.get("responses")) if main_row.get("responses") is not None else {}
                if _extract_status(resp_payload) == "submitted":
                    return json_error("Survey already completed for this household.", 409)
                main_id = main_row["main_questionnaire_id"]
            else:
                draft_payload = json.dumps({
                    "responses": {},
                    "submitted_at": None,
                    "household_id": household_id,
                    "submitted_by_user_id": session.get("user_id"),
                    "submitted_by_username": session.get("username"),
                    "status": "draft"
                })
                main_res = db.session.execute(sql_text("""
                    INSERT INTO main_questionnaire_responses (household_id, user_id, responses)
                    VALUES (:hid, :uid, :resp)
                """), {"hid": household_id, "uid": session.get("user_id"), "resp": draft_payload})
                main_id = main_res.lastrowid

            session["household_id"] = household_id
            session["main_questionnaire_id"] = main_id
            _ensure_survey_contributor(main_id, session.get("user_id"))

            draft_row = db.session.execute(sql_text("""
                SELECT response_data
                FROM household_response_drafts
                WHERE household_id = :hid
                ORDER BY updated_at DESC
                LIMIT 1
            """), {"hid": household_id}).mappings().fetchone()
            draft_payload = safe_json_load(draft_row["response_data"]) if draft_row else {}
            combined = _combine_draft_sections(draft_payload)

            db.session.commit()
            return jsonify({
                "success": True,
                "household_id": household_id,
                "main_questionnaire_id": main_id,
                "resumed": True,
                "draft": draft_payload,
                "responses": combined
            })

        village_id = int(data["village_id"])
        state_id = int(data["state_id"])
        district_id = int(data["district_id"])
        block_id = int(data["block_id"])
        sub_center_id = int(data["sub_center_id"])

        v = db.session.execute(sql_text("""
            SELECT village_id, district_id, block_id, sub_center_id
            FROM villages
            WHERE village_id = :vid
        """), {"vid": village_id}).mappings().fetchone()

        if not v:
            return json_error(f"Village with ID {village_id} does not exist. Contact admin.", 400)

        if int(v["district_id"]) != district_id:
            return json_error("Village does not belong to selected District", 400)
        if v["block_id"] is not None and int(v["block_id"]) != block_id:
            return json_error("Village does not belong to selected Block", 400)
        if v["sub_center_id"] is not None and int(v["sub_center_id"]) != sub_center_id:
            return json_error("Village does not belong to selected Sub-Center", 400)

        insert_res = db.session.execute(sql_text("""
            INSERT INTO households (name, state_id, district_id, block_id, sub_center_id, village_id, user_id)
            VALUES (:name, :state_id, :district_id, :block_id, :sub_center_id, :village_id, :user_id)
        """), {
            "name": data["household_name"],
            "state_id": state_id,
            "district_id": district_id,
            "block_id": block_id,
            "sub_center_id": sub_center_id,
            "village_id": village_id,
            "user_id": session["user_id"],
        })
        household_id = insert_res.lastrowid
        db.session.commit()
        session["household_id"] = household_id

        # Create a draft main questionnaire record immediately for this household
        draft_payload = json.dumps({
            "responses": {},
            "submitted_at": None,
            "household_id": household_id,
            "submitted_by_user_id": session.get("user_id"),
            "submitted_by_username": session.get("username"),
            "status": "draft"
        })
        main_res = db.session.execute(sql_text("""
            INSERT INTO main_questionnaire_responses (household_id, user_id, responses)
            VALUES (:hid, :uid, :resp)
        """), {"hid": household_id, "uid": session.get("user_id"), "resp": draft_payload})
        main_id = main_res.lastrowid
        _ensure_survey_contributor(main_id, session.get("user_id"))
        db.session.commit()
        session["main_questionnaire_id"] = main_id

        return jsonify({"success": True, "household_id": household_id, "main_questionnaire_id": main_id, "resumed": False})

    except Exception as e:
        db.session.rollback()
        return json_error(str(e), 500)


@app.route("/api/resume-household", methods=["POST"])
def resume_household():
    if session.get("role") != "user":
        return json_error("Unauthorized", 401)

    try:
        data = json_body()
        household_name = (data.get("household_name") or "").strip()
        if not household_name:
            return json_error("household_name is required", 400)

        existing = db.session.execute(
            sql_text("SELECT household_id FROM households WHERE LOWER(name) = LOWER(:name) LIMIT 1"),
            {"name": household_name}
        ).mappings().fetchone()
        if not existing:
            return json_error("Household not found", 404)

        household_id = existing["household_id"]
        main_row = db.session.execute(sql_text("""
            SELECT main_questionnaire_id, user_id, responses
            FROM main_questionnaire_responses
            WHERE household_id = :hid
            ORDER BY main_questionnaire_id DESC
            LIMIT 1
        """), {"hid": household_id}).mappings().fetchone()

        if main_row:
            resp_payload = safe_json_load(main_row.get("responses")) if main_row.get("responses") is not None else {}
            if _extract_status(resp_payload) == "submitted":
                return json_error("Survey already completed for this household.", 409)
            main_id = main_row["main_questionnaire_id"]
        else:
            draft_payload = json.dumps({
                "responses": {},
                "submitted_at": None,
                "household_id": household_id,
                "submitted_by_user_id": session.get("user_id"),
                "submitted_by_username": session.get("username"),
                "status": "draft"
            })
            main_res = db.session.execute(sql_text("""
                INSERT INTO main_questionnaire_responses (household_id, user_id, responses)
                VALUES (:hid, :uid, :resp)
            """), {"hid": household_id, "uid": session.get("user_id"), "resp": draft_payload})
            main_id = main_res.lastrowid

        session["household_id"] = household_id
        session["main_questionnaire_id"] = main_id
        _ensure_survey_contributor(main_id, session.get("user_id"))

        draft_row = db.session.execute(sql_text("""
            SELECT response_data
            FROM household_response_drafts
            WHERE household_id = :hid
            ORDER BY updated_at DESC
            LIMIT 1
        """), {"hid": household_id}).mappings().fetchone()
        draft_payload = safe_json_load(draft_row["response_data"]) if draft_row else {}
        combined = _combine_draft_sections(draft_payload)

        db.session.commit()
        return jsonify({
            "success": True,
            "household_id": household_id,
            "main_questionnaire_id": main_id,
            "resumed": True,
            "draft": draft_payload,
            "responses": combined
        })
    except Exception as e:
        db.session.rollback()
        return json_error(str(e), 500)


@app.route("/api/questionnaire-data")
def get_questionnaire_data():
    try:
        sections = db.session.execute(
            sql_text("SELECT * FROM questionnaire_sections ORDER BY section_order")
        ).mappings().all()

        result = []
        for section in sections:
            question_rows = db.session.execute(
                sql_text("SELECT * FROM questions WHERE question_section_id=:sid ORDER BY question_order"),
                {"sid": section["section_id"]},
            ).mappings().all()
            questions = build_question_tree([dict(q) for q in question_rows])

            result.append({
                "section_id": section["section_id"],
                "section_title": section["section_title"],
                "section_order": section["section_order"],
                "questions": questions,
            })

        return jsonify(result)
    except Exception as e:
        return json_error(str(e), 500)

@app.route("/api/individual-questionnaire-data")
def get_individual_questionnaire_data():
    try:
        sections = db.session.execute(
            sql_text("SELECT * FROM individual_questionnaire_sections ORDER BY section_order")
        ).mappings().all()

        result = []
        for section in sections:
            question_rows = db.session.execute(
                sql_text("SELECT * FROM individual_questions WHERE question_section_id=:sid ORDER BY question_order"),
                {"sid": section["section_id"]},
            ).mappings().all()
            questions = build_question_tree([dict(q) for q in question_rows])

            result.append({
                "section_id": section["section_id"],
                "section_title": section["section_title"],
                "section_order": section["section_order"],
                "questions": questions,
            })

        return jsonify(result)
    except Exception as e:
        return json_error(str(e), 500)


@app.route("/api/sections")
def get_sections():
    try:
        rows = db.session.execute(sql_text("""
            SELECT * FROM questionnaire_sections
            ORDER BY section_order
        """)).mappings().all()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        return json_error(str(e), 500)


@app.route("/api/responses", methods=["POST"])
def save_responses():
    if session.get("role") != "user" or "household_id" not in session:
        return json_error("Unauthorized", 401)

    try:
        data = json_body()
        household_id = session["household_id"]
        user_id = session.get("user_id")

        if not data or not data.get("responses"):
            return jsonify({"success": True, "message": "No responses to save"})

        section_id = data.get("section_id")
        if section_id is None:
            return json_error("section_id is required", 400)
        responses = data.get("responses", {})
        timestamp = data.get("timestamp", datetime.now().isoformat())

        household = db.session.execute(
            sql_text("SELECT household_id FROM households WHERE household_id=:hid"),
            {"hid": household_id}
        ).mappings().fetchone()

        if not household:
            return json_error("Household not found", 404)

        row = db.session.execute(sql_text("""
            SELECT draft_id, user_id, response_data
            FROM household_response_drafts
            WHERE household_id = :hid
            ORDER BY updated_at DESC
            LIMIT 1
        """), {"hid": household_id}).mappings().fetchone()

        existing = {}
        if row and row.get("response_data"):
            try:
                existing = json.loads(row["response_data"]) if isinstance(row["response_data"], str) else row["response_data"]
            except Exception:
                existing = {}

        sections = existing.get("sections", {})
        survey_state = existing.get("survey_state")
        sections[str(section_id)] = {
            "responses": responses,
            "timestamp": timestamp
        }

        payload = {
            "sections": sections,
            "last_saved_section_id": section_id,
            "saved_by_user_id": user_id,
            "saved_by_username": session.get("username"),
            "updated_at": timestamp
        }
        if survey_state is not None:
            payload["survey_state"] = survey_state

        if row and row.get("draft_id"):
            db.session.execute(sql_text("""
                UPDATE household_response_drafts
                SET response_data = :data
                WHERE draft_id = :did
            """), {
                "did": row["draft_id"],
                "data": json.dumps(payload)
            })
        else:
            db.session.execute(sql_text("""
                INSERT INTO household_response_drafts (household_id, user_id, response_data)
                VALUES (:hid, :uid, :data)
            """), {
                "hid": household_id,
                "uid": user_id,
                "data": json.dumps(payload)
            })

        main_id = safe_int(session.get("main_questionnaire_id"))
        if not main_id:
            main_row = db.session.execute(sql_text("""
                SELECT main_questionnaire_id
                FROM main_questionnaire_responses
                WHERE household_id = :hid
                ORDER BY main_questionnaire_id DESC
                LIMIT 1
            """), {"hid": household_id}).mappings().fetchone()
            if main_row:
                main_id = main_row["main_questionnaire_id"]
            else:
                draft_payload = json.dumps({
                    "responses": {},
                    "submitted_at": None,
                    "household_id": household_id,
                    "submitted_by_user_id": user_id,
                    "submitted_by_username": session.get("username"),
                    "status": "draft"
                })
                main_res = db.session.execute(sql_text("""
                    INSERT INTO main_questionnaire_responses (household_id, user_id, responses)
                    VALUES (:hid, :uid, :resp)
                """), {"hid": household_id, "uid": user_id, "resp": draft_payload})
                main_id = main_res.lastrowid
            session["main_questionnaire_id"] = main_id

        _ensure_survey_contributor(main_id, user_id)

        db.session.commit()
        return jsonify({"success": True, "message": f"Section {section_id} responses saved", "saved_count": len(responses)})

    except Exception as e:
        db.session.rollback()
        return json_error(str(e), 500)

@app.route("/api/main-questionnaire/submit", methods=["POST"])
def submit_main_questionnaire():
    if session.get("role") != "user" or "household_id" not in session:
        return json_error("Unauthorized", 401)

    try:
        data = json_body()
        responses = data.get("responses", None)

        household_id = safe_int(session.get("household_id"))
        user_id = session.get("user_id")
        if not household_id:
            return json_error("Household not set. Please create/select a household before submitting.", 400)

        household = db.session.execute(
            sql_text("SELECT household_id FROM households WHERE household_id=:hid"),
            {"hid": household_id}
        ).mappings().fetchone()
        if not household:
            return json_error("Household not found. Please create/select a household before submitting.", 400)

        # Fallback: if responses are missing/empty, rebuild from draft sections
        if not isinstance(responses, dict) or len(responses) == 0:
            draft_row = db.session.execute(sql_text("""
                SELECT response_data
                FROM household_response_drafts
                WHERE household_id = :hid
                ORDER BY updated_at DESC
                LIMIT 1
            """), {"hid": household_id}).mappings().fetchone()
            draft_payload = safe_json_load(draft_row["response_data"]) if draft_row else {}
            combined = {}
            for sec in (draft_payload or {}).get("sections", {}).values():
                sec_resp = (sec or {}).get("responses") or {}
                if isinstance(sec_resp, dict):
                    combined.update(sec_resp)
            responses = combined

        if not isinstance(responses, dict) or len(responses) == 0:
            return json_error("responses is required", 400)

        payload = json.dumps({
            "responses": responses,
            "status": "household_submitted",
            "household_submitted_at": datetime.now().isoformat(),
            "household_id": household_id,
            "submitted_by_user_id": user_id,
            "submitted_by_username": session.get("username")
        })

        existing_id = safe_int(data.get("main_questionnaire_id")) or safe_int(session.get("main_questionnaire_id"))
        if not existing_id:
            existing_row = db.session.execute(sql_text("""
                SELECT main_questionnaire_id
                FROM main_questionnaire_responses
                WHERE household_id = :hid
                ORDER BY main_questionnaire_id DESC
                LIMIT 1
            """), {"hid": household_id}).mappings().fetchone()
            if existing_row:
                existing_id = existing_row["main_questionnaire_id"]
        if existing_id:
            row = db.session.execute(sql_text("""
                SELECT main_questionnaire_id
                FROM main_questionnaire_responses
                WHERE main_questionnaire_id = :mid AND household_id = :hid
            """), {"mid": existing_id, "hid": household_id}).mappings().fetchone()
            if row:
                db.session.execute(sql_text("""
                    UPDATE main_questionnaire_responses
                    SET responses = :resp
                    WHERE main_questionnaire_id = :mid
                """), {"resp": payload, "mid": existing_id})
                _ensure_survey_contributor(existing_id, user_id)
                db.session.commit()
                session["main_questionnaire_id"] = existing_id
                return jsonify({"success": True, "main_questionnaire_id": existing_id})
            latest_row = db.session.execute(sql_text("""
                SELECT main_questionnaire_id
                FROM main_questionnaire_responses
                WHERE household_id = :hid
                ORDER BY main_questionnaire_id DESC
                LIMIT 1
            """), {"hid": household_id}).mappings().fetchone()
            if latest_row:
                existing_id = latest_row["main_questionnaire_id"]
                db.session.execute(sql_text("""
                    UPDATE main_questionnaire_responses
                    SET responses = :resp
                    WHERE main_questionnaire_id = :mid
                """), {"resp": payload, "mid": existing_id})
                _ensure_survey_contributor(existing_id, user_id)
                db.session.commit()
                session["main_questionnaire_id"] = existing_id
                return jsonify({"success": True, "main_questionnaire_id": existing_id})

        insert_res = db.session.execute(sql_text("""
            INSERT INTO main_questionnaire_responses (household_id, user_id, responses)
            VALUES (:hid, :uid, :resp)
        """), {"hid": household_id, "uid": user_id, "resp": payload})

        new_id = insert_res.lastrowid
        _ensure_survey_contributor(new_id, user_id)
        db.session.commit()
        session["main_questionnaire_id"] = new_id
        return jsonify({"success": True, "main_questionnaire_id": new_id})

    except Exception as e:
        db.session.rollback()
        return json_error(str(e), 500, traceback.format_exc())

@app.route("/api/complete-household", methods=["POST"])
def complete_household_survey():
    if session.get("role") != "user" or "household_id" not in session:
        return json_error("Unauthorized", 401)

    try:
        data = json_body()
        household_id = safe_int(session.get("household_id"))
        user_id = session.get("user_id")
        if not household_id:
            return json_error("Household not set.", 400)

        main_id = safe_int(data.get("main_questionnaire_id")) or safe_int(session.get("main_questionnaire_id"))
        if not main_id:
            row = db.session.execute(sql_text("""
                SELECT main_questionnaire_id
                FROM main_questionnaire_responses
                WHERE household_id = :hid
                ORDER BY main_questionnaire_id DESC
                LIMIT 1
            """), {"hid": household_id}).mappings().fetchone()
            if row:
                main_id = row["main_questionnaire_id"]
        if not main_id:
            return json_error("Main questionnaire not found.", 404)

        row = db.session.execute(sql_text("""
            SELECT responses
            FROM main_questionnaire_responses
            WHERE main_questionnaire_id = :mid AND household_id = :hid
        """), {"mid": main_id, "hid": household_id}).mappings().fetchone()
        if not row:
            return json_error("Main questionnaire not found.", 404)

        payload = safe_json_load(row["responses"]) or {}
        if _extract_status(payload) != "submitted":
            payload["status"] = "submitted"
            payload["completed_at"] = datetime.now().isoformat()
            db.session.execute(sql_text("""
                UPDATE main_questionnaire_responses
                SET responses = :resp
                WHERE main_questionnaire_id = :mid
            """), {"resp": payload, "mid": main_id})

        _ensure_survey_contributor(main_id, user_id)
        db.session.commit()
        session["main_questionnaire_id"] = main_id
        return jsonify({"success": True, "main_questionnaire_id": main_id})
    except Exception as e:
        db.session.rollback()
        return json_error(str(e), 500)

@app.route("/api/individual-questionnaire/submit", methods=["POST"])
def submit_individual_questionnaire():
    if session.get("role") != "user" or "household_id" not in session:
        return json_error("Unauthorized", 401)

    try:
        data = json_body()
        responses = data.get("responses", None)
        main_id = safe_int(data.get("main_questionnaire_id"))
        member_meta = data.get("member") or {}
        aadhar = (member_meta.get("aadhar") or "").strip()

        if responses is None or not main_id:
            return json_error("main_questionnaire_id and responses are required", 400)
        if not aadhar or not aadhar.isdigit() or len(aadhar) != 12:
            return json_error("Valid 12-digit Aadhaar ID is required", 400)

        aadhar_hash = hash_aadhar(aadhar)
        member_meta["aadhar"] = encrypt_aadhar(aadhar)
        member_meta["aadhar_hash"] = aadhar_hash

        household_id = safe_int(session.get("household_id"))
        user_id = session.get("user_id")
        if not household_id:
            return json_error("Household not set. Please create/select a household before submitting.", 400)

        main_row = db.session.execute(sql_text("""
            SELECT main_questionnaire_id
            FROM main_questionnaire_responses
            WHERE main_questionnaire_id = :mid AND household_id = :hid
        """), {"mid": main_id, "hid": household_id}).mappings().fetchone()

        if not main_row:
            fallback = db.session.execute(sql_text("""
                SELECT main_questionnaire_id
                FROM main_questionnaire_responses
                WHERE household_id = :hid
                ORDER BY main_questionnaire_id DESC
                LIMIT 1
            """), {"hid": household_id}).mappings().fetchone()
            if fallback:
                main_id = fallback["main_questionnaire_id"]
            else:
                # Create a minimal draft main questionnaire so individual responses can be linked
                draft_payload = json.dumps({
                    "responses": {},
                    "submitted_at": None,
                    "household_id": household_id,
                    "submitted_by_user_id": user_id,
                    "submitted_by_username": session.get("username"),
                    "status": "draft"
                })
                insert_main = db.session.execute(sql_text("""
                    INSERT INTO main_questionnaire_responses (household_id, user_id, responses)
                    VALUES (:hid, :uid, :resp)
                """), {"hid": household_id, "uid": user_id, "resp": draft_payload})
                db.session.commit()
                main_id = insert_main.lastrowid
            session["main_questionnaire_id"] = main_id

        payload = json.dumps({
            "member": member_meta,
            "responses": responses,
            "submitted_at": datetime.now().isoformat(),
            "household_id": household_id,
            "submitted_by_user_id": user_id,
            "submitted_by_username": session.get("username")
        })

        insert_ind = db.session.execute(sql_text("""
            INSERT INTO individual_questionnaire_responses (responses, aadhar_hash)
            VALUES (:resp, :ah)
        """), {"resp": payload, "ah": aadhar_hash})
        db.session.commit()

        individual_id = insert_ind.lastrowid

        db.session.execute(sql_text("""
            INSERT INTO main_individual_questionnaire_links
            (main_questionnaire_id, individual_questionnaire_id, household_id, filled_by_user_id)
            VALUES (:mid, :iid, :hid, :uid)
        """), {
            "mid": main_id,
            "iid": individual_id,
            "hid": household_id,
            "uid": user_id
        })
        db.session.commit()

        return jsonify({"success": True, "individual_questionnaire_id": individual_id})

    except Exception as e:
        db.session.rollback()
        return json_error(str(e), 500, traceback.format_exc())

@app.route("/api/admin/individual-lookup")
@role_required("admin")
def admin_individual_lookup():
    aadhar = (request.args.get("aadhar") or "").strip()
    if not aadhar:
        return json_error("aadhar is required", 400)
    aadhar_hash = hash_aadhar(aadhar)

    try:
        row = db.session.execute(sql_text("""
            SELECT ir.individual_questionnaire_id, ir.responses,
                   l.main_questionnaire_id, l.household_id
            FROM individual_questionnaire_responses ir
            JOIN main_individual_questionnaire_links l
              ON l.individual_questionnaire_id = ir.individual_questionnaire_id
            WHERE ir.aadhar_hash = :ah
               OR JSON_UNQUOTE(JSON_EXTRACT(ir.responses, '$.member.aadhar')) = :aadhar
            ORDER BY ir.individual_questionnaire_id DESC
            LIMIT 1
        """), {"aadhar": aadhar, "ah": aadhar_hash}).mappings().fetchone()

        if not row:
            return jsonify({"found": False})

        individual_resp = safe_json_load(row["responses"])
        member = (individual_resp or {}).get("member") or {}
        if member.get("aadhar"):
            member["aadhar"] = decrypt_aadhar(member.get("aadhar"))

        main_row = db.session.execute(sql_text("""
            SELECT main_questionnaire_id, responses
            FROM main_questionnaire_responses
            WHERE main_questionnaire_id = :mid
        """), {"mid": row["main_questionnaire_id"]}).mappings().fetchone()
        main_resp = safe_json_load(main_row["responses"]) if main_row else None

        members_rows = db.session.execute(sql_text("""
            SELECT ir.individual_questionnaire_id, ir.responses
            FROM individual_questionnaire_responses ir
            JOIN main_individual_questionnaire_links l
              ON l.individual_questionnaire_id = ir.individual_questionnaire_id
            WHERE l.household_id = :hid
            ORDER BY ir.individual_questionnaire_id ASC
        """), {"hid": row["household_id"]}).mappings().all()

        household_members = []
        for r in members_rows:
            resp = safe_json_load(r["responses"])
            mem = (resp or {}).get("member") or {}
            if mem.get("aadhar"):
                mem["aadhar"] = decrypt_aadhar(mem.get("aadhar"))
            household_members.append({
                "individual_questionnaire_id": r["individual_questionnaire_id"],
                "member": mem
            })

        return jsonify({
            "found": True,
            "individual_questionnaire_id": row["individual_questionnaire_id"],
            "main_questionnaire_id": row["main_questionnaire_id"],
            "household_id": row["household_id"],
            "member": member,
            "individual_response": individual_resp,
            "main_response": main_resp,
            "household_members": household_members
        })

    except Exception as e:
        return json_error(str(e), 500, traceback.format_exc())


@app.route("/api/admin/individual-lookup/export")
@role_required("admin")
def admin_individual_lookup_export():
    aadhar = (request.args.get("aadhar") or "").strip()
    scope = (request.args.get("scope") or "full").strip().lower()
    if not aadhar:
        return json_error("aadhar is required", 400)
    if scope not in ("full", "main", "individual"):
        return json_error("scope must be full, main, or individual", 400)
    aadhar_hash = hash_aadhar(aadhar)

    try:
        row = db.session.execute(sql_text("""
            SELECT ir.individual_questionnaire_id, ir.responses,
                   l.main_questionnaire_id, l.household_id
            FROM individual_questionnaire_responses ir
            JOIN main_individual_questionnaire_links l
              ON l.individual_questionnaire_id = ir.individual_questionnaire_id
            WHERE ir.aadhar_hash = :ah
               OR JSON_UNQUOTE(JSON_EXTRACT(ir.responses, '$.member.aadhar')) = :aadhar
            ORDER BY ir.individual_questionnaire_id DESC
            LIMIT 1
        """), {"aadhar": aadhar, "ah": aadhar_hash}).mappings().fetchone()

        if not row:
            return json_error("No individual found for this Aadhaar ID", 404)

        individual_resp = safe_json_load(row["responses"]) or {}
        member = (individual_resp or {}).get("member") or {}
        if member.get("aadhar"):
            member["aadhar"] = decrypt_aadhar(member.get("aadhar"))

        main_row = db.session.execute(sql_text("""
            SELECT main_questionnaire_id, responses
            FROM main_questionnaire_responses
            WHERE main_questionnaire_id = :mid
        """), {"mid": row["main_questionnaire_id"]}).mappings().fetchone()
        main_resp = safe_json_load(main_row["responses"]) if main_row else {}

        rows = []
        headers = [
            "survey_type",
            "household_id",
            "main_questionnaire_id",
            "individual_questionnaire_id",
            "member_aadhar",
            "member_name",
            "section_order",
            "section_title",
            "question_order",
            "question_id",
            "question_text",
            "answer",
            "submitted_at",
            "submitted_by",
        ]

        if scope in ("full", "main"):
            main_index = build_question_index("questions", "questionnaire_sections")
            main_payload = main_resp if isinstance(main_resp, dict) else {}
            main_rows = build_export_rows(
                (main_payload or {}).get("responses", {}),
                main_index,
                survey_type="main",
                household_id=row["household_id"],
                main_id=row["main_questionnaire_id"],
                submitted_at=(main_payload or {}).get("submitted_at"),
                submitted_by=(main_payload or {}).get("submitted_by_username"),
                member=member,
            )
            rows.extend(main_rows)

        if scope in ("full", "individual"):
            ind_index = build_question_index("individual_questions", "individual_questionnaire_sections")
            ind_payload = individual_resp if isinstance(individual_resp, dict) else {}
            ind_rows = build_export_rows(
                (ind_payload or {}).get("responses", {}),
                ind_index,
                survey_type="individual",
                household_id=row["household_id"],
                main_id=row["main_questionnaire_id"],
                individual_id=row["individual_questionnaire_id"],
                submitted_at=(ind_payload or {}).get("submitted_at"),
                submitted_by=(ind_payload or {}).get("submitted_by_username"),
                member=member,
            )
            rows.extend(ind_rows)

        filename = f"aadhaar_{aadhar}_{scope}_export.csv"
        return csv_response(filename, rows, headers)

    except Exception as e:
        return json_error(str(e), 500, traceback.format_exc())

@app.route("/api/admin/individual-response/<int:individual_questionnaire_id>")
@role_required("admin")
def admin_get_individual_response(individual_questionnaire_id):
    try:
        row = db.session.execute(sql_text("""
            SELECT individual_questionnaire_id, responses
            FROM individual_questionnaire_responses
            WHERE individual_questionnaire_id = :iid
        """), {"iid": individual_questionnaire_id}).mappings().fetchone()
        if not row:
            return json_error("Individual response not found", 404)

        try:
            payload = json.loads(row["responses"]) if isinstance(row["responses"], str) else row["responses"]
        except Exception:
            payload = row["responses"]

        try:
            member = (payload or {}).get("member") or {}
            if member.get("aadhar"):
                member["aadhar"] = decrypt_aadhar(member.get("aadhar"))
        except Exception:
            pass

        return jsonify({
            "individual_questionnaire_id": row["individual_questionnaire_id"],
            "responses": payload
        })
    except Exception as e:
        return json_error(str(e), 500, traceback.format_exc())


# ==========================================================
# ADMIN APIs (READ)
# ==========================================================

@app.route("/api/admin/sections")
@role_required("admin")
def admin_get_sections():
    try:
        rows = db.session.execute(sql_text("""
            SELECT * FROM questionnaire_sections
            ORDER BY section_order
        """)).mappings().all()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        return json_error(str(e), 500)

@app.route("/api/admin/individual-sections")
@role_required("admin")
def admin_get_individual_sections():
    try:
        rows = db.session.execute(sql_text("""
            SELECT * FROM individual_questionnaire_sections
            ORDER BY section_order
        """)).mappings().all()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        return json_error(str(e), 500)


@app.route("/api/admin/section/<int:section_id>/questions")
@role_required("admin")
def admin_get_section_questions(section_id):
    try:
        rows = db.session.execute(
            sql_text("SELECT * FROM questions WHERE question_section_id=:sid ORDER BY question_order"),
            {"sid": section_id},
        ).mappings().all()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        return json_error(str(e), 500)

@app.route("/api/admin/individual-section/<int:section_id>/questions")
@role_required("admin")
def admin_get_individual_section_questions(section_id):
    try:
        rows = db.session.execute(
            sql_text("SELECT * FROM individual_questions WHERE question_section_id=:sid ORDER BY question_order"),
            {"sid": section_id},
        ).mappings().all()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        return json_error(str(e), 500)


@app.route("/api/admin/users")
@role_required("admin")
def admin_get_users():
    search = (request.args.get("search") or "").strip()
    try:
        if search:
            rows = db.session.execute(
                sql_text("SELECT user_id, username, created_at FROM users WHERE username LIKE :s ORDER BY username"),
                {"s": f"%{search}%"},
            ).mappings().all()
        else:
            rows = db.session.execute(
                sql_text("SELECT user_id, username, created_at FROM users ORDER BY username")
            ).mappings().all()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        try:
            if search:
                rows = db.session.execute(
                    sql_text("SELECT user_id, username FROM users WHERE username LIKE :s ORDER BY username"),
                    {"s": f"%{search}%"},
                ).mappings().all()
            else:
                rows = db.session.execute(
                    sql_text("SELECT user_id, username FROM users ORDER BY username")
                ).mappings().all()
            payload = []
            for r in rows:
                item = dict(r)
                item["created_at"] = None
                payload.append(item)
            return jsonify(payload)
        except Exception:
            return json_error(str(e), 500)


@app.route("/api/admin/admins")
@role_required("admin")
def admin_get_admins():
    try:
        rows = db.session.execute(
            sql_text("SELECT admin_id, username, created_at FROM admins ORDER BY username")
        ).mappings().all()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        try:
            rows = db.session.execute(
                sql_text("SELECT admin_id, username FROM admins ORDER BY username")
            ).mappings().all()
            payload = []
            for r in rows:
                item = dict(r)
                item["created_at"] = None
                payload.append(item)
            return jsonify(payload)
        except Exception:
            return json_error(str(e), 500)


@app.route("/api/admin/all-households")
@role_required("admin")
def admin_get_all_households():
    try:
        search = (request.args.get("search") or "").strip()
        base = """
            SELECT
                h.household_id,
                h.name AS head_name,
                COALESCE(u.username, u_owner.username, '—') AS username,
                s.name AS state_name,
                d.name AS district_name,
                b.name AS block_name,
                sc.name AS sub_center_name,
                v.village_lgd_code AS village_lgd_code
            FROM households h
            LEFT JOIN (
                SELECT household_id, MAX(main_questionnaire_id) AS main_questionnaire_id
                FROM main_questionnaire_responses
                GROUP BY household_id
            ) mm ON mm.household_id = h.household_id
            LEFT JOIN survey_contributors contrib ON contrib.main_questionnaire_id = mm.main_questionnaire_id
            LEFT JOIN users u ON contrib.user_id = u.user_id
            LEFT JOIN users u_owner ON h.user_id = u_owner.user_id
            LEFT JOIN states s ON h.state_id = s.state_id
            LEFT JOIN districts d ON h.district_id = d.district_id
            LEFT JOIN blocks b ON h.block_id = b.block_id
            LEFT JOIN sub_centers sc ON h.sub_center_id = sc.sub_center_id
            LEFT JOIN villages v ON h.village_id = v.village_id
        """
        params = {}
        if search:
            base += " WHERE h.name LIKE :q OR u.username LIKE :q OR u_owner.username LIKE :q OR CAST(v.village_lgd_code AS CHAR) LIKE :q OR CAST(h.household_id AS CHAR) LIKE :q"
            params["q"] = f"%{search}%"
        base += " ORDER BY h.household_id DESC LIMIT 500"

        rows = db.session.execute(sql_text(base), params).mappings().all()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        return json_error(str(e), 500)

@app.route("/api/admin/households/filter")
@role_required("admin")
def admin_filter_households():
    try:
        search = (request.args.get("q") or "").strip()
        state_id = safe_int(request.args.get("state_id"))
        district_id = safe_int(request.args.get("district_id"))
        block_id = safe_int(request.args.get("block_id"))
        sub_center_id = safe_int(request.args.get("sub_center_id"))
        village_id = safe_int(request.args.get("village_id"))

        base = """
            SELECT
                h.household_id,
                h.name AS head_name,
                COALESCE(u.username, u_owner.username, 'â€”') AS username,
                s.name AS state_name,
                d.name AS district_name,
                b.name AS block_name,
                sc.name AS sub_center_name,
                v.village_name,
                v.village_lgd_code,
                m.main_questionnaire_id,
                m.responses
            FROM households h
            INNER JOIN users u ON h.user_id = u.user_id
            INNER JOIN states s ON h.state_id = s.state_id
            INNER JOIN districts d ON h.district_id = d.district_id
            INNER JOIN blocks b ON h.block_id = b.block_id
            INNER JOIN sub_centers sc ON h.sub_center_id = sc.sub_center_id
            INNER JOIN villages v ON h.village_id = v.village_id
            INNER JOIN (
                SELECT household_id, MAX(main_questionnaire_id) AS main_questionnaire_id
                FROM main_questionnaire_responses
                GROUP BY household_id
            ) mm ON mm.household_id = h.household_id
            INNER JOIN main_questionnaire_responses m ON m.main_questionnaire_id = mm.main_questionnaire_id
        """

        filters = []
        params = {}

        if search:
            filters.append("(h.name LIKE :q OR u.username LIKE :q OR u_owner.username LIKE :q OR CAST(v.village_lgd_code AS CHAR) LIKE :q OR CAST(h.household_id AS CHAR) LIKE :q)")
            params["q"] = f"%{search}%"
        if state_id:
            filters.append("h.state_id = :state_id")
            params["state_id"] = state_id
        if district_id:
            filters.append("h.district_id = :district_id")
            params["district_id"] = district_id
        if block_id:
            filters.append("h.block_id = :block_id")
            params["block_id"] = block_id
        if sub_center_id:
            filters.append("h.sub_center_id = :sub_center_id")
            params["sub_center_id"] = sub_center_id
        if village_id:
            filters.append("h.village_id = :village_id")
            params["village_id"] = village_id

        if filters:
            base += " WHERE " + " AND ".join(filters)

        base += " ORDER BY h.household_id DESC LIMIT 500"

        rows = db.session.execute(sql_text(base), params).mappings().all()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        return json_error(str(e), 500)


@app.route("/api/admin/households/count")
@role_required("admin")
def admin_count_households():
    try:
        state_id = safe_int(request.args.get("state_id"))
        district_id = safe_int(request.args.get("district_id"))
        block_id = safe_int(request.args.get("block_id"))
        sub_center_id = safe_int(request.args.get("sub_center_id"))
        village_id = safe_int(request.args.get("village_id"))

        base = """
            SELECT COUNT(*) AS cnt
            FROM households h
        """
        filters = []
        params = {}
        if state_id:
            filters.append("h.state_id = :state_id")
            params["state_id"] = state_id
        if district_id:
            filters.append("h.district_id = :district_id")
            params["district_id"] = district_id
        if block_id:
            filters.append("h.block_id = :block_id")
            params["block_id"] = block_id
        if sub_center_id:
            filters.append("h.sub_center_id = :sub_center_id")
            params["sub_center_id"] = sub_center_id
        if village_id:
            filters.append("h.village_id = :village_id")
            params["village_id"] = village_id

        if filters:
            base += " WHERE " + " AND ".join(filters)

        row = db.session.execute(sql_text(base), params).mappings().fetchone()
        count = int(row["cnt"]) if row and "cnt" in row else 0
        return jsonify({"total_households": count})
    except Exception as e:
        return json_error(str(e), 500, traceback.format_exc())


@app.route("/api/admin/households/export")
@role_required("admin")
def admin_export_household_responses():
    try:
        state_id = safe_int(request.args.get("state_id"))
        district_id = safe_int(request.args.get("district_id"))
        block_id = safe_int(request.args.get("block_id"))
        sub_center_id = safe_int(request.args.get("sub_center_id"))
        village_id = safe_int(request.args.get("village_id"))

        base = """
            SELECT
                h.household_id,
                TRIM(h.name) AS head_name,
                COALESCE(TRIM(u.username), '-') AS username,
                TRIM(s.name) AS state_name,
                TRIM(d.name) AS district_name,
                TRIM(b.name) AS block_name,
                TRIM(sc.name) AS sub_center_name,
                TRIM(v.name) AS village_name,
                v.village_lgd_code AS village_lgd_code,
                m.main_questionnaire_id,
                m.responses
            FROM households h
            INNER JOIN users u ON h.user_id = u.user_id
            INNER JOIN states s ON h.state_id = s.state_id
            INNER JOIN districts d ON h.district_id = d.district_id
            INNER JOIN blocks b ON h.block_id = b.block_id
            INNER JOIN sub_centers sc ON h.sub_center_id = sc.sub_center_id
            INNER JOIN villages v ON h.village_id = v.village_id
            INNER JOIN (
                SELECT household_id, MAX(main_questionnaire_id) AS main_questionnaire_id
                FROM main_questionnaire_responses
                GROUP BY household_id
            ) mm ON mm.household_id = h.household_id
            INNER JOIN main_questionnaire_responses m ON m.main_questionnaire_id = mm.main_questionnaire_id
        """

        filters = []
        params = {}
        if state_id:
            filters.append("h.state_id = :state_id")
            params["state_id"] = state_id
        if district_id:
            filters.append("h.district_id = :district_id")
            params["district_id"] = district_id
        if block_id:
            filters.append("h.block_id = :block_id")
            params["block_id"] = block_id
        if sub_center_id:
            filters.append("h.sub_center_id = :sub_center_id")
            params["sub_center_id"] = sub_center_id
        if village_id:
            filters.append("h.village_id = :village_id")
            params["village_id"] = village_id
        if filters:
            base += " WHERE " + " AND ".join(filters)
        base += " ORDER BY h.household_id DESC"

        rows = db.session.execute(sql_text(base), params).mappings().all()

        question_index = build_question_index("questions", "questionnaire_sections")
        headers = [
            "survey_type",
            "household_id",
            "head_name",
            "username",
            "state_name",
            "district_name",
            "block_name",
            "sub_center_name",
            "village_name",
            "village_lgd_code",
            "main_questionnaire_id",
            "individual_questionnaire_id",
            "member_name",
            "member_aadhar",
            "submitted_at",
            "submitted_by",
        ]

        out_rows = []
        for r in rows:
            payload = safe_json_load(r["responses"]) if r.get("responses") else {}
            main_rows = build_export_rows(
                (payload or {}).get("responses", {}),
                question_index,
                survey_type="main",
                household_id=r.get("household_id"),
                main_id=r.get("main_questionnaire_id"),
                submitted_at=(payload or {}).get("submitted_at"),
                submitted_by=(payload or {}).get("submitted_by_username"),
            )
            if not main_rows:
                continue
            extra = {
                "head_name": r.get("head_name") or "",
                "username": r.get("username") or "",
                "state_name": r.get("state_name") or "",
                "district_name": r.get("district_name") or "",
                "block_name": r.get("block_name") or "",
                "sub_center_name": r.get("sub_center_name") or "",
                "village_name": r.get("village_name") or "",
                "village_lgd_code": r.get("village_lgd_code") or "",
            }
            for row in main_rows:
                row.update(extra)
            out_rows.extend(main_rows)

        filename = "household_responses_export.csv"
        return csv_response(filename, out_rows, headers)
    except Exception as e:
        return json_error(str(e), 500, traceback.format_exc())


@app.route("/api/admin/households/export/main-wide")
@role_required("admin")
def admin_export_main_wide():
    try:
        state_id = safe_int(request.args.get("state_id"))
        district_id = safe_int(request.args.get("district_id"))
        block_id = safe_int(request.args.get("block_id"))
        sub_center_id = safe_int(request.args.get("sub_center_id"))
        village_id = safe_int(request.args.get("village_id"))

        base = """
            SELECT
                h.household_id,
                TRIM(h.name) AS head_name,
                COALESCE(TRIM(u.username), '-') AS username,
                TRIM(s.name) AS state_name,
                TRIM(d.name) AS district_name,
                TRIM(b.name) AS block_name,
                TRIM(sc.name) AS sub_center_name,
                TRIM(v.name) AS village_name,
                v.village_lgd_code AS village_lgd_code,
                m.main_questionnaire_id,
                m.responses
            FROM households h
            INNER JOIN users u ON h.user_id = u.user_id
            INNER JOIN states s ON h.state_id = s.state_id
            INNER JOIN districts d ON h.district_id = d.district_id
            INNER JOIN blocks b ON h.block_id = b.block_id
            INNER JOIN sub_centers sc ON h.sub_center_id = sc.sub_center_id
            INNER JOIN villages v ON h.village_id = v.village_id
            INNER JOIN (
                SELECT household_id, MAX(main_questionnaire_id) AS main_questionnaire_id
                FROM main_questionnaire_responses
                GROUP BY household_id
            ) mm ON mm.household_id = h.household_id
            INNER JOIN main_questionnaire_responses m ON m.main_questionnaire_id = mm.main_questionnaire_id
        """
        filters = []
        params = {}
        if state_id:
            filters.append("h.state_id = :state_id")
            params["state_id"] = state_id
        if district_id:
            filters.append("h.district_id = :district_id")
            params["district_id"] = district_id
        if block_id:
            filters.append("h.block_id = :block_id")
            params["block_id"] = block_id
        if sub_center_id:
            filters.append("h.sub_center_id = :sub_center_id")
            params["sub_center_id"] = sub_center_id
        if village_id:
            filters.append("h.village_id = :village_id")
            params["village_id"] = village_id
        if filters:
            base += " WHERE " + " AND ".join(filters)
        base += " ORDER BY h.household_id DESC"

        rows = db.session.execute(sql_text(base), params).mappings().all()

        question_index = build_question_index("questions", "questionnaire_sections")
        q_headers, q_map = build_wide_headers(question_index)

        headers = [
            "household_id",
            "head_name",
            "username",
            "state_name",
            "district_name",
            "block_name",
            "sub_center_name",
            "village_name",
            "village_lgd_code",
            "main_questionnaire_id",
            "individual_questionnaire_id",
            "member_name",
            "member_aadhar",
            "submitted_at",
            "submitted_by",
        ] + q_headers

        out_rows = []
        for r in rows:
            payload = safe_json_load(r["responses"]) if r.get("responses") else {}
            base_row = {
                "household_id": r.get("household_id") or "",
                "head_name": r.get("head_name") or "",
                "username": r.get("username") or "",
                "state_name": r.get("state_name") or "",
                "district_name": r.get("district_name") or "",
                "block_name": r.get("block_name") or "",
                "sub_center_name": r.get("sub_center_name") or "",
                "village_name": r.get("village_name") or "",
                "village_lgd_code": r.get("village_lgd_code") or "",
                "main_questionnaire_id": r.get("main_questionnaire_id") or "",
                "individual_questionnaire_id": r.get("individual_questionnaire_id") or "",
                "member_name": " ".join([member.get("first_name", ""), member.get("middle_name", ""), member.get("surname", "")]).strip(),
                "member_aadhar": member.get("aadhar") or "",
                "submitted_at": (payload or {}).get("submitted_at") or "",
                "submitted_by": (payload or {}).get("submitted_by_username") or "",
            }
            wide = build_wide_row((payload or {}).get("responses", {}), q_map)
            base_row.update(wide)
            out_rows.append(base_row)

        return csv_response("main_questionnaire_wide.csv", out_rows, headers)
    except Exception as e:
        return json_error(str(e), 500, traceback.format_exc())


@app.route("/api/admin/households/export/individual-wide")
@role_required("admin")
def admin_export_individual_wide():
    try:
        state_id = safe_int(request.args.get("state_id"))
        district_id = safe_int(request.args.get("district_id"))
        block_id = safe_int(request.args.get("block_id"))
        sub_center_id = safe_int(request.args.get("sub_center_id"))
        village_id = safe_int(request.args.get("village_id"))

        base = """
            SELECT
                h.household_id,
                TRIM(h.name) AS head_name,
                COALESCE(TRIM(u.username), '-') AS username,
                TRIM(s.name) AS state_name,
                TRIM(d.name) AS district_name,
                TRIM(b.name) AS block_name,
                TRIM(sc.name) AS sub_center_name,
                TRIM(v.name) AS village_name,
                v.village_lgd_code AS village_lgd_code,
                l.main_questionnaire_id,
                ir.individual_questionnaire_id,
                ir.responses
            FROM individual_questionnaire_responses ir
            INNER JOIN main_individual_questionnaire_links l
                ON l.individual_questionnaire_id = ir.individual_questionnaire_id
            INNER JOIN households h ON h.household_id = l.household_id
            INNER JOIN users u ON h.user_id = u.user_id
            INNER JOIN states s ON h.state_id = s.state_id
            INNER JOIN districts d ON h.district_id = d.district_id
            INNER JOIN blocks b ON h.block_id = b.block_id
            INNER JOIN sub_centers sc ON h.sub_center_id = sc.sub_center_id
            INNER JOIN villages v ON h.village_id = v.village_id
        """
        filters = []
        params = {}
        if state_id:
            filters.append("h.state_id = :state_id")
            params["state_id"] = state_id
        if district_id:
            filters.append("h.district_id = :district_id")
            params["district_id"] = district_id
        if block_id:
            filters.append("h.block_id = :block_id")
            params["block_id"] = block_id
        if sub_center_id:
            filters.append("h.sub_center_id = :sub_center_id")
            params["sub_center_id"] = sub_center_id
        if village_id:
            filters.append("h.village_id = :village_id")
            params["village_id"] = village_id
        if filters:
            base += " WHERE " + " AND ".join(filters)
        base += " ORDER BY ir.individual_questionnaire_id DESC"

        rows = db.session.execute(sql_text(base), params).mappings().all()

        question_index = build_question_index("individual_questions", "individual_questionnaire_sections")
        q_headers, q_map = build_wide_headers(question_index)

        headers = [
            "household_id",
            "head_name",
            "username",
            "state_name",
            "district_name",
            "block_name",
            "sub_center_name",
            "village_name",
            "village_lgd_code",
            "main_questionnaire_id",
            "individual_questionnaire_id",
            "member_name",
            "member_aadhar",
            "submitted_at",
            "submitted_by",
        ] + q_headers

        out_rows = []
        for r in rows:
            payload = safe_json_load(r["responses"]) if r.get("responses") else {}
            member = (payload or {}).get("member") or {}
            if member.get("aadhar"):
                try:
                    member["aadhar"] = decrypt_aadhar(member.get("aadhar"))
                except Exception:
                    pass
            member_name = " ".join([member.get("first_name", ""), member.get("middle_name", ""), member.get("surname", "")]).strip()
            base_row = {
                "household_id": r.get("household_id") or "",
                "head_name": r.get("head_name") or "",
                "username": r.get("username") or "",
                "state_name": r.get("state_name") or "",
                "district_name": r.get("district_name") or "",
                "block_name": r.get("block_name") or "",
                "sub_center_name": r.get("sub_center_name") or "",
                "village_name": r.get("village_name") or "",
                "village_lgd_code": r.get("village_lgd_code") or "",
                "main_questionnaire_id": r.get("main_questionnaire_id") or "",
                "individual_questionnaire_id": r.get("individual_questionnaire_id") or "",
                "member_name": member_name,
                "member_aadhar": member.get("aadhar") or "",
                "submitted_at": (payload or {}).get("submitted_at") or "",
                "submitted_by": (payload or {}).get("submitted_by_username") or "",
            }
            wide = build_wide_row((payload or {}).get("responses", {}), q_map)
            base_row.update(wide)
            out_rows.append(base_row)

        return csv_response("individual_questionnaire_wide.csv", out_rows, headers)
    except Exception as e:
        return json_error(str(e), 500, traceback.format_exc())


@app.route("/api/admin/households/export/count")
@role_required("admin")
def admin_export_household_responses_count():
    try:
        state_id = safe_int(request.args.get("state_id"))
        district_id = safe_int(request.args.get("district_id"))
        block_id = safe_int(request.args.get("block_id"))
        sub_center_id = safe_int(request.args.get("sub_center_id"))
        village_id = safe_int(request.args.get("village_id"))

        base = """
            SELECT COUNT(DISTINCT h.household_id) AS cnt
            FROM households h
            INNER JOIN users u ON h.user_id = u.user_id
            INNER JOIN states s ON h.state_id = s.state_id
            INNER JOIN districts d ON h.district_id = d.district_id
            INNER JOIN blocks b ON h.block_id = b.block_id
            INNER JOIN sub_centers sc ON h.sub_center_id = sc.sub_center_id
            INNER JOIN villages v ON h.village_id = v.village_id
            INNER JOIN (
                SELECT household_id, MAX(main_questionnaire_id) AS main_questionnaire_id
                FROM main_questionnaire_responses
                GROUP BY household_id
            ) mm ON mm.household_id = h.household_id
            INNER JOIN main_questionnaire_responses m ON m.main_questionnaire_id = mm.main_questionnaire_id
        """
        filters = []
        params = {}
        if state_id:
            filters.append("h.state_id = :state_id")
            params["state_id"] = state_id
        if district_id:
            filters.append("h.district_id = :district_id")
            params["district_id"] = district_id
        if block_id:
            filters.append("h.block_id = :block_id")
            params["block_id"] = block_id
        if sub_center_id:
            filters.append("h.sub_center_id = :sub_center_id")
            params["sub_center_id"] = sub_center_id
        if village_id:
            filters.append("h.village_id = :village_id")
            params["village_id"] = village_id

        if filters:
            base += " WHERE " + " AND ".join(filters)

        row = db.session.execute(sql_text(base), params).mappings().fetchone()
        count = int(row["cnt"]) if row and "cnt" in row else 0
        return jsonify({"total_households": count})
    except Exception as e:
        return json_error(str(e), 500, traceback.format_exc())


@app.route("/api/admin/search-households")
@role_required("admin")
def admin_search_households():
    q = (request.args.get("q") or "").strip()
    if not q:
        return jsonify([])
    params = {"q": f"%{q}%"}
    rows = db.session.execute(sql_text("""
        SELECT
            h.household_id,
            h.name AS head_name,
            COALESCE(u.username, u_owner.username, '—') AS username,
            v.village_lgd_code AS village_lgd_code
        FROM households h
        LEFT JOIN users u ON u.user_id = h.user_id
        LEFT JOIN villages v ON v.village_id = h.village_id
        WHERE h.name LIKE :q
           OR CAST(v.village_lgd_code AS CHAR) LIKE :q
           OR u.username LIKE :q
           OR CAST(h.household_id AS CHAR) LIKE :q
        ORDER BY h.household_id DESC
        LIMIT 500
    """), params).mappings().all()
    return jsonify([dict(r) for r in rows])


@app.route("/api/admin/household/<int:household_id>", methods=["DELETE"])
@role_required("admin")
def admin_delete_household(household_id):
    try:
        def table_exists(name: str) -> bool:
            row = db.session.execute(sql_text("""
                SELECT 1
                FROM information_schema.tables
                WHERE table_schema = DATABASE() AND table_name = :t
                LIMIT 1
            """), {"t": name}).mappings().fetchone()
            return bool(row)

        def table_has_column(name: str, col: str) -> bool:
            row = db.session.execute(sql_text("""
                SELECT 1
                FROM information_schema.columns
                WHERE table_schema = DATABASE()
                  AND table_name = :t
                  AND column_name = :c
                LIMIT 1
            """), {"t": name, "c": col}).mappings().fetchone()
            return bool(row)

        exists = db.session.execute(
            sql_text("SELECT household_id FROM households WHERE household_id = :hid"),
            {"hid": household_id},
        ).mappings().fetchone()
        if not exists:
            return json_error("Household not found", 404)

        # Delete dependent records in safe order, but only if tables/columns exist
        if table_exists("household_response_drafts"):
            db.session.execute(
                sql_text("DELETE FROM household_response_drafts WHERE household_id = :hid"),
                {"hid": household_id},
            )

        if table_exists("main_individual_questionnaire_links"):
            if table_exists("individual_questionnaire_responses"):
                db.session.execute(sql_text("""
                    DELETE FROM individual_questionnaire_responses
                    WHERE individual_questionnaire_id IN (
                        SELECT individual_questionnaire_id
                        FROM main_individual_questionnaire_links
                        WHERE household_id = :hid
                    )
                """), {"hid": household_id})

            db.session.execute(
                sql_text("DELETE FROM main_individual_questionnaire_links WHERE household_id = :hid"),
                {"hid": household_id},
            )

        if table_exists("main_questionnaire_responses"):
            db.session.execute(
                sql_text("DELETE FROM main_questionnaire_responses WHERE household_id = :hid"),
                {"hid": household_id},
            )

        # Clean up person/survey_attempt tables if present
        for person_table in ["persons", "person"]:
            if table_exists(person_table) and table_has_column(person_table, "household_id"):
                for attempt_table in ["survey_attempts", "survey_attempt"]:
                    if table_exists(attempt_table) and table_has_column(attempt_table, "person_id"):
                        db.session.execute(sql_text(f"""
                            DELETE FROM {attempt_table}
                            WHERE person_id IN (
                                SELECT person_id FROM {person_table} WHERE household_id = :hid
                            )
                        """), {"hid": household_id})
                db.session.execute(
                    sql_text(f"DELETE FROM {person_table} WHERE household_id = :hid"),
                    {"hid": household_id},
                )

        db.session.execute(
            sql_text("DELETE FROM households WHERE household_id = :hid"),
            {"hid": household_id},
        )

        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return json_error(str(e), 500, traceback.format_exc())


@app.route("/api/admin/search-users")
@role_required("admin")
def admin_search_users():
    q = (request.args.get("q") or "").strip()
    if not q:
        return jsonify([])
    params = {"q": f"%{q}%"}
    try:
        rows = db.session.execute(sql_text("""
            SELECT user_id, username, created_at
            FROM users
            WHERE username LIKE :q
            ORDER BY username
        """), params).mappings().all()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        try:
            rows = db.session.execute(sql_text("""
                SELECT user_id, username
                FROM users
                WHERE username LIKE :q
                ORDER BY username
            """), params).mappings().all()
            payload = []
            for r in rows:
                item = dict(r)
                item["created_at"] = None
                payload.append(item)
            return jsonify(payload)
        except Exception:
            return json_error(str(e), 500)


# ==========================================================
# ADMIN APIs (WRITE) - SECTIONS
# ==========================================================

@app.route("/api/admin/section", methods=["POST"])
@role_required("admin")
def admin_create_section():
    try:
        data = json_body()
        title = (data.get("section_title") or data.get("title") or "").strip()
        if not title:
            return json_error("section_title is required", 400)

        next_order = db.session.execute(sql_text("""
            SELECT COALESCE(MAX(section_order), 0) + 1 AS next_order
            FROM questionnaire_sections
        """)).mappings().fetchone()["next_order"]

        db.session.execute(sql_text("""
            INSERT INTO questionnaire_sections (section_title, section_order)
            VALUES (:t, :o)
        """), {"t": title, "o": next_order})
        db.session.commit()

        new_id = db.session.execute(sql_text("SELECT LAST_INSERT_ID() AS id")).mappings().fetchone()["id"]
        return jsonify({"success": True, "section_id": new_id, "section_order": next_order})

    except Exception as e:
        db.session.rollback()
        return json_error(str(e), 500, traceback.format_exc())


@app.route("/api/admin/section/<int:section_id>", methods=["PUT"])
@role_required("admin")
def admin_update_section(section_id):
    try:
        data = json_body()
        title = (data.get("section_title") or data.get("title") or "").strip()
        if not title:
            return json_error("section_title is required", 400)

        db.session.execute(sql_text("""
            UPDATE questionnaire_sections
            SET section_title = :t
            WHERE section_id = :id
        """), {"t": title, "id": section_id})
        db.session.commit()
        return jsonify({"success": True})

    except Exception as e:
        db.session.rollback()
        return json_error(str(e), 500, traceback.format_exc())


@app.route("/api/admin/section/<int:section_id>", methods=["DELETE"])
@role_required("admin")
def admin_delete_section(section_id):
    try:
        # if you want cascade, ensure FK constraints in DB. Otherwise delete questions first:
        db.session.execute(sql_text("DELETE FROM questions WHERE question_section_id = :sid"), {"sid": section_id})
        db.session.execute(sql_text("DELETE FROM questionnaire_sections WHERE section_id = :sid"), {"sid": section_id})
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return json_error(str(e), 500, traceback.format_exc())


# ==========================================================
# ADMIN APIs (WRITE) - LOCATIONS
# ==========================================================

@app.route("/api/admin/state", methods=["POST"])
@role_required("admin")
def admin_create_state():
    try:
        data = json_body()
        name = (data.get("name") or "").strip()
        if not name:
            return json_error("name is required", 400)

        columns = []
        params = {}

        # state_id may not be auto-increment in some schemas
        if column_exists("states", "state_id") and not column_is_auto_increment("states", "state_id"):
            columns.append("state_id")
            params["state_id"] = next_id("states", "state_id")

        columns.append("name")
        params["name"] = name

        if column_exists("states", "territory_type"):
            territory_type = (data.get("territory_type") or "STATE").strip()
            columns.append("territory_type")
            params["territory_type"] = territory_type

        cols_sql = ", ".join(columns)
        vals_sql = ", ".join(f":{c}" for c in columns)
        db.session.execute(sql_text(f"""
            INSERT INTO states ({cols_sql})
            VALUES ({vals_sql})
        """), params)
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        print("=== ERROR ADDING VILLAGE ===")
        print(traceback.format_exc())
        return json_error(str(e), 500, traceback.format_exc())


@app.route("/api/admin/state/<int:state_id>", methods=["PUT"])
@role_required("admin")
def admin_update_state(state_id):
    try:
        data = json_body()
        name = (data.get("name") or "").strip()
        if not name:
            return json_error("name is required", 400)

        if column_exists("states", "territory_type"):
            territory_type = (data.get("territory_type") or "STATE").strip()
            db.session.execute(sql_text("""
                UPDATE states
                SET name = :name,
                    territory_type = :territory_type
                WHERE state_id = :id
            """), {"name": name, "territory_type": territory_type, "id": state_id})
        else:
            db.session.execute(sql_text("""
                UPDATE states
                SET name = :name
                WHERE state_id = :id
            """), {"name": name, "id": state_id})
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return json_error(str(e), 500, traceback.format_exc())

@app.route("/api/admin/individual-section/<int:section_id>", methods=["DELETE"])
@role_required("admin")
def admin_delete_individual_section(section_id):
    try:
        db.session.execute(sql_text("DELETE FROM individual_questions WHERE question_section_id = :sid"), {"sid": section_id})
        db.session.execute(sql_text("DELETE FROM individual_questionnaire_sections WHERE section_id = :sid"), {"sid": section_id})
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return json_error(str(e), 500, traceback.format_exc())

@app.route("/api/admin/individual-section/<int:section_id>", methods=["PUT"])
@role_required("admin")
def admin_update_individual_section(section_id):
    try:
        data = json_body()
        title = (data.get("section_title") or data.get("title") or "").strip()
        if not title:
            return json_error("section_title is required", 400)

        db.session.execute(sql_text("""
            UPDATE individual_questionnaire_sections
            SET section_title = :t
            WHERE section_id = :id
        """), {"t": title, "id": section_id})
        db.session.commit()
        return jsonify({"success": True})

    except Exception as e:
        db.session.rollback()
        return json_error(str(e), 500, traceback.format_exc())

@app.route("/api/admin/individual-section", methods=["POST"])
@role_required("admin")
def admin_create_individual_section():
    try:
        data = json_body()
        title = (data.get("section_title") or data.get("title") or "").strip()
        if not title:
            return json_error("section_title is required", 400)

        next_order = db.session.execute(sql_text("""
            SELECT COALESCE(MAX(section_order), 0) + 1 AS next_order
            FROM individual_questionnaire_sections
        """)).mappings().fetchone()["next_order"]

        db.session.execute(sql_text("""
            INSERT INTO individual_questionnaire_sections (section_title, section_order)
            VALUES (:t, :o)
        """), {"t": title, "o": next_order})
        db.session.commit()

        new_id = db.session.execute(sql_text("SELECT LAST_INSERT_ID() AS id")).mappings().fetchone()["id"]
        return jsonify({"success": True, "section_id": new_id, "section_order": next_order})

    except Exception as e:
        db.session.rollback()
        return json_error(str(e), 500, traceback.format_exc())


@app.route("/api/admin/state/<int:state_id>", methods=["DELETE"])
@role_required("admin")
def admin_delete_state(state_id):
    try:
        db.session.execute(sql_text("DELETE FROM states WHERE state_id = :id"), {"id": state_id})
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return json_error(str(e), 500, traceback.format_exc())


@app.route("/api/admin/district", methods=["POST"])
@role_required("admin")
def admin_create_district():
    try:
        data = json_body()
        name = (data.get("name") or "").strip()
        state_id = safe_int(data.get("state_id"))
        if not name or not state_id:
            return json_error("name and state_id are required", 400)

        columns = []
        params = {"name": name, "state_id": state_id}

        if column_exists("districts", "district_id") and not column_is_auto_increment("districts", "district_id"):
            columns.append("district_id")
            params["district_id"] = next_id("districts", "district_id")

        columns += ["name", "state_id"]
        cols_sql = ", ".join(columns)
        vals_sql = ", ".join(f":{c}" for c in columns)
        db.session.execute(sql_text(f"""
            INSERT INTO districts ({cols_sql})
            VALUES ({vals_sql})
        """), params)
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return json_error(str(e), 500, traceback.format_exc())


@app.route("/api/admin/district/<int:district_id>", methods=["PUT"])
@role_required("admin")
def admin_update_district(district_id):
    try:
        data = json_body()
        name = (data.get("name") or "").strip()
        state_id = safe_int(data.get("state_id"))
        if not name or not state_id:
            return json_error("name and state_id are required", 400)

        db.session.execute(sql_text("""
            UPDATE districts
            SET name = :name,
                state_id = :state_id
            WHERE district_id = :id
        """), {"name": name, "state_id": state_id, "id": district_id})
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return json_error(str(e), 500, traceback.format_exc())


@app.route("/api/admin/district/<int:district_id>", methods=["DELETE"])
@role_required("admin")
def admin_delete_district(district_id):
    try:
        db.session.execute(sql_text("DELETE FROM districts WHERE district_id = :id"), {"id": district_id})
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return json_error(str(e), 500, traceback.format_exc())


@app.route("/api/admin/block", methods=["POST"])
@role_required("admin")
def admin_create_block():
    try:
        data = json_body()
        name = (data.get("name") or "").strip()
        district_id = safe_int(data.get("district_id"))
        if not name or not district_id:
            return json_error("name and district_id are required", 400)

        columns = []
        params = {"name": name, "district_id": district_id}

        if column_exists("blocks", "block_id") and not column_is_auto_increment("blocks", "block_id"):
            columns.append("block_id")
            params["block_id"] = next_id("blocks", "block_id")

        columns += ["name", "district_id"]
        cols_sql = ", ".join(columns)
        vals_sql = ", ".join(f":{c}" for c in columns)
        db.session.execute(sql_text(f"""
            INSERT INTO blocks ({cols_sql})
            VALUES ({vals_sql})
        """), params)
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return json_error(str(e), 500, traceback.format_exc())


@app.route("/api/admin/block/<int:block_id>", methods=["PUT"])
@role_required("admin")
def admin_update_block(block_id):
    try:
        data = json_body()
        name = (data.get("name") or "").strip()
        district_id = safe_int(data.get("district_id"))
        if not name or not district_id:
            return json_error("name and district_id are required", 400)

        db.session.execute(sql_text("""
            UPDATE blocks
            SET name = :name,
                district_id = :district_id
            WHERE block_id = :id
        """), {"name": name, "district_id": district_id, "id": block_id})
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return json_error(str(e), 500, traceback.format_exc())


@app.route("/api/admin/block/<int:block_id>", methods=["DELETE"])
@role_required("admin")
def admin_delete_block(block_id):
    try:
        db.session.execute(sql_text("DELETE FROM blocks WHERE block_id = :id"), {"id": block_id})
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return json_error(str(e), 500, traceback.format_exc())


@app.route("/api/admin/sub-center", methods=["POST"])
@role_required("admin")
def admin_create_sub_center():
    try:
        data = json_body()
        name = (data.get("name") or "").strip()
        block_id = safe_int(data.get("block_id"))
        if not name or not block_id:
            return json_error("name and block_id are required", 400)

        columns = []
        params = {"name": name, "block_id": block_id}

        if column_exists("sub_centers", "sub_center_id") and not column_is_auto_increment("sub_centers", "sub_center_id"):
            columns.append("sub_center_id")
            params["sub_center_id"] = next_id("sub_centers", "sub_center_id")

        columns += ["name", "block_id"]
        cols_sql = ", ".join(columns)
        vals_sql = ", ".join(f":{c}" for c in columns)
        db.session.execute(sql_text(f"""
            INSERT INTO sub_centers ({cols_sql})
            VALUES ({vals_sql})
        """), params)
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return json_error(str(e), 500, traceback.format_exc())


@app.route("/api/admin/sub-center/<int:sub_center_id>", methods=["PUT"])
@role_required("admin")
def admin_update_sub_center(sub_center_id):
    try:
        data = json_body()
        name = (data.get("name") or "").strip()
        block_id = safe_int(data.get("block_id"))
        if not name or not block_id:
            return json_error("name and block_id are required", 400)

        db.session.execute(sql_text("""
            UPDATE sub_centers
            SET name = :name,
                block_id = :block_id
            WHERE sub_center_id = :id
        """), {"name": name, "block_id": block_id, "id": sub_center_id})
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return json_error(str(e), 500, traceback.format_exc())


@app.route("/api/admin/sub-center/<int:sub_center_id>", methods=["DELETE"])
@role_required("admin")
def admin_delete_sub_center(sub_center_id):
    try:
        db.session.execute(sql_text("DELETE FROM sub_centers WHERE sub_center_id = :id"), {"id": sub_center_id})
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return json_error(str(e), 500, traceback.format_exc())


@app.route("/api/admin/village", methods=["POST"])
@role_required("admin")
def admin_create_village():
    try:
        data = json_body()
        lgd = safe_int(data.get("village_lgd_code"))
        name = (data.get("name") or "").strip()
        district_id = safe_int(data.get("district_id"))
        block_id = safe_int(data.get("block_id"))
        sub_center_id = safe_int(data.get("sub_center_id"))

        has_sub_center = column_exists("villages", "sub_center_id")
        if not lgd or not name or not district_id or not block_id or (has_sub_center and not sub_center_id):
            return json_error("village_lgd_code, name, district_id, block_id, sub_center_id are required", 400)

        existing = db.session.execute(sql_text("""
            SELECT village_lgd_code
            FROM villages
            WHERE village_lgd_code = :lgd
            LIMIT 1
        """), {"lgd": lgd}).mappings().fetchone()
        if existing:
            return json_error("Village LGD code already exists. Use a unique LGD code.", 409)

        columns = []
        params = {
            "lgd": lgd,
            "name": name,
            "district_id": district_id,
            "block_id": block_id,
            "sub_center_id": sub_center_id
        }

        if column_exists("villages", "village_id") and not column_is_auto_increment("villages", "village_id"):
            columns.append("village_id")
            params["village_id"] = next_id("villages", "village_id")

        if column_exists("villages", "village_lgd_code"):
            columns.append("village_lgd_code")
            params["village_lgd_code"] = lgd_code

        columns += ["name", "district_id", "block_id"]
        if has_sub_center:
            columns.append("sub_center_id")

        cols_sql = ", ".join(columns)
        vals_sql = ", ".join(f":{c}" for c in columns)
        db.session.execute(sql_text(f"""
            INSERT INTO villages ({cols_sql})
            VALUES ({vals_sql})
        """), params)
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return json_error(str(e), 500, traceback.format_exc())


@app.route("/api/admin/village/<int:village_lgd_code>", methods=["PUT"])
@role_required("admin")
def admin_update_village(village_lgd_code):
    try:
        data = json_body()
        name = (data.get("name") or "").strip()
        district_id = safe_int(data.get("district_id"))
        block_id = safe_int(data.get("block_id"))
        sub_center_id = safe_int(data.get("sub_center_id"))
        has_sub_center = column_exists("villages", "sub_center_id")

        if not name or not district_id or not block_id or (has_sub_center and not sub_center_id):
            return json_error("name, district_id, block_id, sub_center_id are required", 400)

        if has_sub_center:
            db.session.execute(sql_text("""
                UPDATE villages
                SET name = :name,
                    district_id = :district_id,
                    block_id = :block_id,
                    sub_center_id = :sub_center_id
                WHERE village_lgd_code = :lgd
            """), {
                "name": name,
                "district_id": district_id,
                "block_id": block_id,
                "sub_center_id": sub_center_id,
                "lgd": village_lgd_code
            })
        else:
            db.session.execute(sql_text("""
                UPDATE villages
                SET name = :name,
                    district_id = :district_id,
                    block_id = :block_id
                WHERE village_lgd_code = :lgd
            """), {
                "name": name,
                "district_id": district_id,
                "block_id": block_id,
                "lgd": village_lgd_code
            })
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return json_error(str(e), 500, traceback.format_exc())


@app.route("/api/admin/village/<int:village_lgd_code>", methods=["DELETE"])
@role_required("admin")
def admin_delete_village(village_lgd_code):
    try:
        db.session.execute(sql_text("DELETE FROM villages WHERE village_lgd_code = :lgd"), {"lgd": village_lgd_code})
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return json_error(str(e), 500, traceback.format_exc())


@app.route("/api/admin/locations/bulk/<level>", methods=["POST"])
@role_required("admin")
def admin_bulk_locations(level):
    if not load_workbook:
        return json_error("openpyxl is required for bulk upload. Please install it on the server.", 500)

    level_map = {
        "state": "states",
        "states": "states",
        "district": "districts",
        "districts": "districts",
        "block": "blocks",
        "blocks": "blocks",
        "subcenter": "subcenters",
        "subcentre": "subcenters",
        "subcenters": "subcenters",
        "subcentres": "subcenters",
        "village": "villages",
        "villages": "villages",
    }
    level = level_map.get((level or "").strip().lower())
    if not level:
        return json_error("Invalid level", 400)

    file = request.files.get("file")
    if not file:
        return json_error("file is required", 400)

    try:
        rows = read_excel_rows(file)
    except Exception as e:
        return json_error(f"Failed to read Excel: {str(e)}", 400)

    if not rows:
        return json_error("No data found in file", 400)

    items, parse_errors = parse_bulk_rows(rows, level)
    if not items and parse_errors:
        return json_error("No valid rows found", 400, {"errors": parse_errors[:20]})

    inserted = 0
    skipped = 0
    errors = []

    def get_state_id(name: str):
        row = db.session.execute(
            sql_text("SELECT state_id FROM states WHERE LOWER(name) = LOWER(:n) LIMIT 1"),
            {"n": name},
        ).mappings().fetchone()
        if row:
            return row["state_id"], False

        columns = []
        params = {}
        if column_exists("states", "state_id") and not column_is_auto_increment("states", "state_id"):
            columns.append("state_id")
            params["state_id"] = next_id("states", "state_id")

        columns.append("name")
        params["name"] = name

        if column_exists("states", "territory_type"):
            columns.append("territory_type")
            params["territory_type"] = "STATE"

        cols_sql = ", ".join(columns)
        vals_sql = ", ".join(f":{c}" for c in columns)
        res = db.session.execute(sql_text(f"""
            INSERT INTO states ({cols_sql})
            VALUES ({vals_sql})
        """), params)
        new_id = res.lastrowid or params.get("state_id")
        return new_id, True

    def get_district_id(state_id: int, name: str):
        row = db.session.execute(
            sql_text("SELECT district_id FROM districts WHERE LOWER(name) = LOWER(:n) AND state_id = :sid LIMIT 1"),
            {"n": name, "sid": state_id},
        ).mappings().fetchone()
        if row:
            return row["district_id"], False

        columns = []
        params = {"name": name, "state_id": state_id}
        if column_exists("districts", "district_id") and not column_is_auto_increment("districts", "district_id"):
            columns.append("district_id")
            params["district_id"] = next_id("districts", "district_id")

        columns += ["name", "state_id"]
        cols_sql = ", ".join(columns)
        vals_sql = ", ".join(f":{c}" for c in columns)
        res = db.session.execute(sql_text(f"""
            INSERT INTO districts ({cols_sql})
            VALUES ({vals_sql})
        """), params)
        new_id = res.lastrowid or params.get("district_id")
        return new_id, True

    def get_block_id(district_id: int, name: str):
        row = db.session.execute(
            sql_text("SELECT block_id FROM blocks WHERE LOWER(name) = LOWER(:n) AND district_id = :did LIMIT 1"),
            {"n": name, "did": district_id},
        ).mappings().fetchone()
        if row:
            return row["block_id"], False

        columns = []
        params = {"name": name, "district_id": district_id}
        if column_exists("blocks", "block_id") and not column_is_auto_increment("blocks", "block_id"):
            columns.append("block_id")
            params["block_id"] = next_id("blocks", "block_id")

        columns += ["name", "district_id"]
        cols_sql = ", ".join(columns)
        vals_sql = ", ".join(f":{c}" for c in columns)
        res = db.session.execute(sql_text(f"""
            INSERT INTO blocks ({cols_sql})
            VALUES ({vals_sql})
        """), params)
        new_id = res.lastrowid or params.get("block_id")
        return new_id, True

    def get_subcenter_id(block_id: int, name: str):
        row = db.session.execute(
            sql_text("SELECT sub_center_id FROM sub_centers WHERE LOWER(name) = LOWER(:n) AND block_id = :bid LIMIT 1"),
            {"n": name, "bid": block_id},
        ).mappings().fetchone()
        if row:
            return row["sub_center_id"], False

        columns = []
        params = {"name": name, "block_id": block_id}
        if column_exists("sub_centers", "sub_center_id") and not column_is_auto_increment("sub_centers", "sub_center_id"):
            columns.append("sub_center_id")
            params["sub_center_id"] = next_id("sub_centers", "sub_center_id")

        columns += ["name", "block_id"]
        cols_sql = ", ".join(columns)
        vals_sql = ", ".join(f":{c}" for c in columns)
        res = db.session.execute(sql_text(f"""
            INSERT INTO sub_centers ({cols_sql})
            VALUES ({vals_sql})
        """), params)
        new_id = res.lastrowid or params.get("sub_center_id")
        return new_id, True

    def get_village_id(name: str, district_id: int, block_id: int, sub_center_id: int, lgd_code):
        has_sub_center = column_exists("villages", "sub_center_id")
        params = {"n": name, "did": district_id, "bid": block_id}
        where = "LOWER(name) = LOWER(:n) AND district_id = :did AND block_id = :bid"
        if has_sub_center:
            where += " AND sub_center_id = :sid"
            params["sid"] = sub_center_id

        row = db.session.execute(
            sql_text(f"SELECT village_id, village_lgd_code FROM villages WHERE {where} LIMIT 1"),
            params,
        ).mappings().fetchone()
        if row:
            return row["village_id"], False

        if lgd_code and column_exists("villages", "village_lgd_code"):
            lgd_row = db.session.execute(
                sql_text("SELECT village_id FROM villages WHERE village_lgd_code = :lgd LIMIT 1"),
                {"lgd": lgd_code},
            ).mappings().fetchone()
            if lgd_row:
                return lgd_row["village_id"], False

        if not lgd_code and column_exists("villages", "village_lgd_code"):
            lgd_code = next_id("villages", "village_lgd_code")

        columns = []
        params = {"name": name, "district_id": district_id, "block_id": block_id}
        if has_sub_center:
            params["sub_center_id"] = sub_center_id

        if column_exists("villages", "village_id") and not column_is_auto_increment("villages", "village_id"):
            columns.append("village_id")
            params["village_id"] = next_id("villages", "village_id")

        if column_exists("villages", "village_lgd_code"):
            columns.append("village_lgd_code")
            params["village_lgd_code"] = lgd_code

        columns += ["name", "district_id", "block_id"]
        if has_sub_center:
            columns.append("sub_center_id")

        cols_sql = ", ".join(columns)
        vals_sql = ", ".join(f":{c}" for c in columns)
        res = db.session.execute(sql_text(f"""
            INSERT INTO villages ({cols_sql})
            VALUES ({vals_sql})
        """), params)
        new_id = res.lastrowid or params.get("village_id")
        return new_id, True

    for item in items:
        try:
            with db.session.begin_nested():
                state_id, created = get_state_id(item["state"])

                if level == "states":
                    if created:
                        inserted += 1
                    else:
                        skipped += 1
                    continue

                district_id, created = get_district_id(state_id, item["district"])
                if level == "districts":
                    if created:
                        inserted += 1
                    else:
                        skipped += 1
                    continue

                block_id, created = get_block_id(district_id, item["block"])
                if level == "blocks":
                    if created:
                        inserted += 1
                    else:
                        skipped += 1
                    continue

                sub_center_id, created = get_subcenter_id(block_id, item["subcenter"])
                if level == "subcenters":
                    if created:
                        inserted += 1
                    else:
                        skipped += 1
                    continue

                lgd_code = safe_int(item.get("lgd"))
                village_id, created = get_village_id(
                    item["village"], district_id, block_id, sub_center_id, lgd_code
                )
                if created:
                    inserted += 1
                else:
                    skipped += 1
        except Exception as e:
            errors.append(f"Row {item.get('_row')}: {str(e)}")

    db.session.commit()
    if errors:
        return jsonify({
            "success": True,
            "inserted": inserted,
            "skipped": skipped,
            "errors": errors[:20],
            "parse_errors": parse_errors[:20]
        }), 200

    return jsonify({"success": True, "inserted": inserted, "skipped": skipped, "parse_errors": parse_errors[:20]})

# ==========================================================
# ADMIN APIs (WRITE) - QUESTIONS  ✅ FIXED
# ==========================================================

@app.route("/api/admin/question", methods=["POST"])
@role_required("admin")
def admin_create_question():
    try:
        data = json_body()
        payload = normalize_question_payload(data)

        if not payload["section_id"] or not payload["question_text"] or not payload["question_type"]:
            return json_error("section_id, question_text, question_type are required", 400)

        # Next order must be globally unique (questions.uk_question_order)
        next_order = db.session.execute(sql_text("""
            SELECT COALESCE(MAX(question_order), 0) + 1 AS next_order
            FROM questions
        """)).mappings().fetchone()["next_order"]

        db.session.execute(sql_text("""
            INSERT INTO questions
            (question_section_id, question_text, question_type, answer_type, options, parent_id, trigger_value, question_order)
            VALUES
            (:sid, :qt, :qtype, :atype, :opts, :pid, :tval, :ord)
        """), {
            "sid": payload["section_id"],
            "qt": payload["question_text"],
            "qtype": payload["question_type"],
            "atype": payload["answer_type"],
            "opts": payload["options"],
            "pid": payload["parent_id"],
            "tval": payload["trigger_value"],
            "ord": next_order
        })

        db.session.commit()
        new_id = db.session.execute(sql_text("SELECT LAST_INSERT_ID() AS id")).mappings().fetchone()["id"]
        return jsonify({"success": True, "question_id": new_id, "question_order": next_order})

    except Exception as e:
        db.session.rollback()
        print("=== ERROR ADDING QUESTION ===")
        print(traceback.format_exc())
        return json_error(str(e), 500, traceback.format_exc())

@app.route("/api/admin/individual-question", methods=["POST"])
@role_required("admin")
def admin_create_individual_question():
    try:
        data = json_body()
        payload = normalize_question_payload(data)

        if not payload["section_id"] or not payload["question_text"] or not payload["question_type"]:
            return json_error("section_id, question_text, question_type are required", 400)

        next_order = db.session.execute(sql_text("""
            SELECT COALESCE(MAX(question_order), 0) + 1 AS next_order
            FROM individual_questions
        """)).mappings().fetchone()["next_order"]

        db.session.execute(sql_text("""
            INSERT INTO individual_questions
            (question_section_id, question_text, question_type, answer_type, options, parent_id, trigger_value, question_order)
            VALUES
            (:sid, :qt, :qtype, :atype, :opts, :pid, :tval, :ord)
        """), {
            "sid": payload["section_id"],
            "qt": payload["question_text"],
            "qtype": payload["question_type"],
            "atype": payload["answer_type"],
            "opts": payload["options"],
            "pid": payload["parent_id"],
            "tval": payload["trigger_value"],
            "ord": next_order
        })

        db.session.commit()
        new_id = db.session.execute(sql_text("SELECT LAST_INSERT_ID() AS id")).mappings().fetchone()["id"]
        return jsonify({"success": True, "question_id": new_id, "question_order": next_order})

    except Exception as e:
        db.session.rollback()
        return json_error(str(e), 500, traceback.format_exc())


@app.route("/api/admin/question/<int:question_id>", methods=["PUT"])
@role_required("admin")
def admin_update_question(question_id):
    try:
        data = json_body()
        payload = normalize_question_payload(data)

        if not payload["section_id"] or not payload["question_text"] or not payload["question_type"]:
            return json_error("section_id, question_text, question_type are required", 400)

        db.session.execute(sql_text("""
            UPDATE questions
            SET question_section_id = :sid,
                question_text = :qt,
                question_type = :qtype,
                answer_type = :atype,
                options = :opts,
                parent_id = :pid,
                trigger_value = :tval
            WHERE question_id = :id
        """), {
            "sid": payload["section_id"],
            "qt": payload["question_text"],
            "qtype": payload["question_type"],
            "atype": payload["answer_type"],
            "opts": payload["options"],
            "pid": payload["parent_id"],
            "tval": payload["trigger_value"],
            "id": question_id
        })

        db.session.commit()
        return jsonify({"success": True})

    except Exception as e:
        db.session.rollback()
        return json_error(str(e), 500, traceback.format_exc())

@app.route("/api/admin/individual-question/<int:question_id>", methods=["PUT"])
@role_required("admin")
def admin_update_individual_question(question_id):
    try:
        data = json_body()
        payload = normalize_question_payload(data)

        if not payload["section_id"] or not payload["question_text"] or not payload["question_type"]:
            return json_error("section_id, question_text, question_type are required", 400)

        db.session.execute(sql_text("""
            UPDATE individual_questions
            SET question_section_id = :sid,
                question_text = :qt,
                question_type = :qtype,
                answer_type = :atype,
                options = :opts,
                parent_id = :pid,
                trigger_value = :tval
            WHERE question_id = :id
        """), {
            "sid": payload["section_id"],
            "qt": payload["question_text"],
            "qtype": payload["question_type"],
            "atype": payload["answer_type"],
            "opts": payload["options"],
            "pid": payload["parent_id"],
            "tval": payload["trigger_value"],
            "id": question_id
        })

        db.session.commit()
        return jsonify({"success": True})

    except Exception as e:
        db.session.rollback()
        return json_error(str(e), 500, traceback.format_exc())


@app.route("/api/admin/question/<int:question_id>", methods=["DELETE"])
@role_required("admin")
def admin_delete_question(question_id):
    try:
        db.session.execute(sql_text("DELETE FROM questions WHERE question_id = :id"), {"id": question_id})
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return json_error(str(e), 500, traceback.format_exc())

@app.route("/api/admin/individual-question/<int:question_id>", methods=["DELETE"])
@role_required("admin")
def admin_delete_individual_question(question_id):
    try:
        db.session.execute(sql_text("DELETE FROM individual_questions WHERE question_id = :id"), {"id": question_id})
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return json_error(str(e), 500, traceback.format_exc())


# ==========================================================
# QUESTION TREE (used by questionnaire_manager.html)
# ==========================================================

@app.route("/api/questions/tree/<int:section_id>")
def get_questions_tree(section_id):
    try:
        rows = db.session.execute(sql_text("""
            SELECT question_id, question_section_id, question_text, question_type,
                   answer_type, options, parent_id, trigger_value, question_order
            FROM questions
            WHERE question_section_id = :sid
            ORDER BY parent_id, question_order
        """), {"sid": section_id}).mappings().all()

        questions = [dict(r) for r in rows]

        def build_tree(parent_id=None):
            nodes = []
            for q in questions:
                if (parent_id is None and q.get("parent_id") is None) or (parent_id is not None and q.get("parent_id") == parent_id):
                    node = {
                        "question_id": q["question_id"],
                        "question_text": q["question_text"],
                        "question_type": q["question_type"],
                        "answer_type": q["answer_type"],
                        "options": q.get("options"),
                        "parent_id": q.get("parent_id"),
                        "trigger_value": q.get("trigger_value"),
                        "children": build_tree(q["question_id"]),
                    }
                    nodes.append(node)
            return nodes

        return jsonify(build_tree())
    except Exception as e:
        return json_error(str(e), 500, traceback.format_exc())


# ==========================================================
# ADMIN USER/ADMIN MGMT
# ==========================================================

@app.route("/api/admin/user", methods=["POST"])
@role_required("admin")
def admin_create_user():
    try:
        data = json_body()
        username = (data.get("username") or "").strip()
        password = data.get("password") or ""

        if not username or not password:
            return json_error("username and password are required", 400)
        if len(password) < 6:
            return json_error("Password must be at least 6 characters", 400)

        password_hash = generate_password_hash(password)
        db.session.execute(
            sql_text("INSERT INTO users (username, password_hash) VALUES (:u, :p)"),
            {"u": username, "p": password_hash},
        )
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return json_error(str(e), 500, traceback.format_exc())


@app.route("/api/admin/user/<int:user_id>", methods=["DELETE"])
@role_required("admin")
def admin_delete_user(user_id):
    try:
        db.session.execute(sql_text("DELETE FROM users WHERE user_id=:id"), {"id": user_id})
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return json_error(str(e), 500, traceback.format_exc())


@app.route("/api/admin/admin", methods=["POST"])
@role_required("admin")
def admin_create_admin():
    try:
        data = json_body()
        username = (data.get("username") or "").strip()
        password = data.get("password") or ""

        if not username or not password:
            return json_error("username and password are required", 400)
        if len(password) < 6:
            return json_error("Password must be at least 6 characters", 400)

        password_hash = generate_password_hash(password)
        # admins table requires both password and password_hash (password is NOT NULL)
        db.session.execute(
            sql_text("INSERT INTO admins (username, password, password_hash) VALUES (:u, :p_raw, :p_hash)"),
            {"u": username, "p_raw": password_hash, "p_hash": password_hash},
        )
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return json_error(str(e), 500, traceback.format_exc())


@app.route("/api/admin/admin/<int:admin_id>", methods=["DELETE"])
@role_required("admin")
def admin_delete_admin(admin_id):
    try:
        db.session.execute(sql_text("DELETE FROM admins WHERE admin_id=:id"), {"id": admin_id})
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return json_error(str(e), 500, traceback.format_exc())


# ==========================================================
# DB HEALTH
# ==========================================================

@app.route("/api/health-db")
def health_db():
    try:
        counts = {}
        for table in ["states", "districts", "blocks", "sub_centers", "villages", "households"]:
            r = db.session.execute(sql_text(f"SELECT COUNT(*) as count FROM {table}")).mappings().fetchone()
            counts[table] = int(r["count"]) if r and "count" in r else 0
        return jsonify({"ok": True, "counts": counts})
    except Exception as e:
        return json_error(str(e), 500, traceback.format_exc())


# ==========================================================
# GEOGRAPHIC DATA ENDPOINTS (kept for compatibility)
# ==========================================================

@app.route("/api/admin/all-states")
@role_required("admin")
def admin_get_all_states():
    try:
        search = request.args.get("search", "").strip()
        if search:
            result = db.session.execute(
                sql_text("SELECT * FROM states WHERE name LIKE :search ORDER BY name"),
                {"search": f"%{search}%"},
            ).mappings().all()
        else:
            result = db.session.execute(sql_text("SELECT * FROM states ORDER BY name")).mappings().all()
        return jsonify([dict(r) for r in result])
    except Exception as e:
        return json_error(str(e), 500)


@app.route("/api/admin/all-districts")
@role_required("admin")
def admin_get_all_districts():
    try:
        search = request.args.get("search", "").strip()
        state_id = request.args.get("state_id", "").strip()

        query = "SELECT d.*, s.name as state_name FROM districts d JOIN states s ON d.state_id = s.state_id WHERE 1=1"
        params = {}

        if state_id:
            query += " AND d.state_id = :state_id"
            params["state_id"] = int(state_id)

        if search:
            query += " AND d.name LIKE :search"
            params["search"] = f"%{search}%"

        query += " ORDER BY s.name, d.name"
        result = db.session.execute(sql_text(query), params).mappings().all()
        return jsonify([dict(r) for r in result])
    except Exception as e:
        return json_error(str(e), 500)


@app.route("/api/admin/all-blocks")
@role_required("admin")
def admin_get_all_blocks():
    try:
        search = request.args.get("search", "").strip()
        district_id = request.args.get("district_id", "").strip()

        query = "SELECT b.*, d.name as district_name, s.name as state_name FROM blocks b JOIN districts d ON b.district_id = d.district_id JOIN states s ON d.state_id = s.state_id WHERE 1=1"
        params = {}

        if district_id:
            query += " AND b.district_id = :district_id"
            params["district_id"] = int(district_id)

        if search:
            query += " AND b.name LIKE :search"
            params["search"] = f"%{search}%"

        query += " ORDER BY s.name, d.name, b.name"
        result = db.session.execute(sql_text(query), params).mappings().all()
        return jsonify([dict(r) for r in result])
    except Exception as e:
        return json_error(str(e), 500)


@app.route("/api/admin/all-sub-centers")
@role_required("admin")
def admin_get_all_sub_centers():
    try:
        search = request.args.get("search", "").strip()
        block_id = request.args.get("block_id", "").strip()

        query = """
            SELECT sc.*, b.name as block_name, d.name as district_name, s.name as state_name
            FROM sub_centers sc
            LEFT JOIN blocks b ON sc.block_id = b.block_id
            LEFT JOIN districts d ON b.district_id = d.district_id
            LEFT JOIN states s ON d.state_id = s.state_id
            WHERE 1=1
        """
        params = {}

        if block_id:
            query += " AND sc.block_id = :block_id"
            params["block_id"] = int(block_id)

        if search:
            query += " AND sc.name LIKE :search"
            params["search"] = f"%{search}%"

        query += " ORDER BY s.name, d.name, b.name, sc.name"
        result = db.session.execute(sql_text(query), params).mappings().all()
        return jsonify([dict(r) for r in result])
    except Exception as e:
        return json_error(str(e), 500)


@app.route("/api/admin/all-villages")
@role_required("admin")
def admin_get_all_villages():
    try:
        search = request.args.get("search", "").strip()
        district_id = request.args.get("district_id", "").strip()
        sub_center_id = request.args.get("sub_center_id", "").strip()

        query = """SELECT v.village_id, v.village_lgd_code, v.name, v.district_id, v.block_id, v.sub_center_id,
                   sc.name as sub_center_name, b.name as block_name,
                   d.name as district_name, s.name as state_name
            FROM villages v
            LEFT JOIN sub_centers sc ON v.sub_center_id = sc.sub_center_id
            LEFT JOIN blocks b ON v.block_id = b.block_id
            LEFT JOIN districts d ON v.district_id = d.district_id
            LEFT JOIN states s ON d.state_id = s.state_id
            WHERE 1=1"""
        params = {}

        if sub_center_id:
            query += " AND v.sub_center_id = :sub_center_id"
            params["sub_center_id"] = int(sub_center_id)

        if district_id:
            query += " AND v.district_id = :district_id"
            params["district_id"] = int(district_id)

        if search:
            query += " AND v.name LIKE :search"
            params["search"] = f"%{search}%"

        query += " ORDER BY s.name, d.name, b.name, v.name"
        result = db.session.execute(sql_text(query), params).mappings().all()
        return jsonify([dict(r) for r in result])
    except Exception as e:
        return json_error(str(e), 500)


# ==========================================================
# LOCATIONS API (for survey flow - geographic hierarchy)
# ==========================================================

@app.route("/api/locations/<location_type>/<int:parent_id>", methods=["GET"])
def get_locations(location_type, parent_id):
    """
    Return geographic locations based on type and parent ID.
    Supports: states, districts, blocks, subcentres, villages, households
    """
    try:
        if location_type == "states":
            result = db.session.execute(
                sql_text("SELECT state_id as id, state_name as name FROM states ORDER BY state_name")
            ).mappings().all()
        
        elif location_type == "districts":
            result = db.session.execute(
                sql_text("SELECT district_id as id, district_name as name FROM districts WHERE state_id = :sid ORDER BY district_name"),
                {"sid": parent_id}
            ).mappings().all()
        
        elif location_type == "blocks":
            result = db.session.execute(
                sql_text("SELECT block_id as id, block_name as name FROM blocks WHERE district_id = :did ORDER BY block_name"),
                {"did": parent_id}
            ).mappings().all()
        
        elif location_type == "subcentres":
            result = db.session.execute(
                sql_text("SELECT sub_center_id as id, sub_center_name as name FROM sub_centers WHERE block_id = :bid ORDER BY sub_center_name"),
                {"bid": parent_id}
            ).mappings().all()
        
        elif location_type == "villages":
            result = db.session.execute(
                sql_text("SELECT village_id as id, village_name as name FROM villages WHERE sub_center_id = :sid ORDER BY village_name"),
                {"sid": parent_id}
            ).mappings().all()
        
        elif location_type == "households":
            result = db.session.execute(
                sql_text("SELECT household_id as id, household_name as name FROM households WHERE village_id = :vid ORDER BY household_name"),
                {"vid": parent_id}
            ).mappings().all()
        
        else:
            return json_error(f"Unknown location type: {location_type}", 400)
        
        return jsonify([{"id": r["id"], "name": r["name"]} for r in result])
    
    except Exception as e:
        return json_error(str(e), 500)


@app.route("/api/initialize-survey", methods=["POST"])
def initialize_survey():
    """
    Initialize a survey attempt for a person/household.
    Checks if person exists, creates attempt, etc.
    """
    try:
        data = json_body()
        aadhar = (data.get("aadhar") or "").strip()
        household_id = safe_int(data.get("household_id"))
        age = safe_int(data.get("age"))
        
        if not aadhar or len(aadhar) != 12:
            return json_error("Valid Aadhaar (12 digits) required", 400)
        if not household_id or age is None or age < 0:
            return json_error("Valid household_id and age required", 400)
        
        aadhar_hash = hash_aadhar(aadhar)
        aadhar_enc = encrypt_aadhar(aadhar)

        # Check if person exists
        person = db.session.execute(
            sql_text("SELECT person_id FROM persons WHERE aadhar_hash = :ah AND household_id = :hid"),
            {"ah": aadhar_hash, "hid": household_id}
        ).mappings().fetchone()
        
        person_id = None
        if person:
            person_id = person["person_id"]
            # Check if already completed
            completed = db.session.execute(
                sql_text("SELECT survey_attempt_id FROM survey_attempts WHERE person_id = :pid AND status = 'completed'"),
                {"pid": person_id}
            ).mappings().fetchone()
            
            if completed:
                return jsonify({
                    "status": "fully_completed",
                    "message": "This person's survey is already completed."
                }), 200
        else:
            # Create new person
            db.session.execute(
                sql_text("""
                    INSERT INTO persons (aadhar, aadhar_hash, household_id, age)
                    VALUES (:aadhar, :ah, :hid, :age)
                """),
                {"aadhar": aadhar_enc, "ah": aadhar_hash, "hid": household_id, "age": age}
            )
            db.session.commit()
            
            # Get the newly created person
            person = db.session.execute(
                sql_text("SELECT person_id FROM persons WHERE aadhar_hash = :ah AND household_id = :hid"),
                {"ah": aadhar_hash, "hid": household_id}
            ).mappings().fetchone()
            person_id = person["person_id"]
        
        # Create survey attempt
        db.session.execute(
            sql_text("""
                INSERT INTO survey_attempts (person_id, status, created_at)
                VALUES (:pid, 'in_progress', NOW())
            """),
            {"pid": person_id}
        )
        db.session.commit()
        
        # Get the attempt ID
        attempt = db.session.execute(
            sql_text("SELECT survey_attempt_id FROM survey_attempts WHERE person_id = :pid ORDER BY created_at DESC LIMIT 1"),
            {"pid": person_id}
        ).mappings().fetchone()
        
        return jsonify({
            "status": "ready",
            "person_id": person_id,
            "attempt_id": attempt["survey_attempt_id"],
            "responses": {}
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print("ERROR initializing survey:", traceback.format_exc())
        return json_error(str(e), 500)


# ==========================================================
# RUN
# ==========================================================

# @app.route("/__init_db__", methods=["POST"])
# def init_db():
#     try:
#         with app.app_context():
#             ensure_questionnaire_tables()
#             ensure_household_name_unique_index()
#             ensure_aadhar_storage_columns()
#         return "✅ Database schema created successfully", 200
#     except Exception:
#         return traceback.format_exc(), 500

if __name__ == "__main__":
    app.run(debug=True)
